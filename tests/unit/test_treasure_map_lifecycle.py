from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from dof.api import MAIN_SHEET_NAME, create_or_update_treasure_map


def _write(p: Path, content: bytes) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(content)


def _headers(ws) -> dict[str, int]:
    """Return mapping header -> 1-based column index."""
    hdr: dict[str, int] = {}
    for i, c in enumerate(ws[1], start=1):
        if c.value:
            hdr[str(c.value)] = i
    return hdr


def _rows_by_location(ws, loc_col: int) -> dict[str, int]:
    """Return mapping location string -> row index."""
    out: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, loc_col).value
        if v:
            out[str(v)] = r
    return out


def _cell_date(ws, row: int, col: int) -> date:
    """Return cell value as a date (asserts value is present)."""
    v = ws.cell(row, col).value
    assert v is not None, f"Expected date in r{row}c{col}, got None"
    return v.date()


def test_treasure_map_full_lifecycle_create_update_delete(tmp_path: Path) -> None:
    """
    Lifecycle:
    1) New treasuremap creation
    2) Update preserves Date Found + Description, updates Last Seen, bumps Version on change
    3) Deletion removed from map when prune_missing=True
    Also checks:
    - When prune_missing=False and a file is missing, its row remains AND Last Seen does not advance.
    """
    root = tmp_path / "root"
    _write(root / "a.pdf", b"%PDF-1.4\nhello\n")
    _write(root / "sub" / "b.xlsx", b"fake-xlsx-bytes")
    _write(root / "sub" / "c.text", b"hi")

    out = tmp_path / "treasure_map.xlsx"

    # --- 1) Create new map ----------------------------------------------------
    d1 = date(2025, 12, 18)
    create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        sharepoint_base_url="https://sp.example/doclib",
        today=d1,
    )

    wb1 = load_workbook(out)
    ws1 = wb1[MAIN_SHEET_NAME]
    h1 = _headers(ws1)

    # Required headers present (including new Last Seen)
    assert [ws1.cell(1, i).value for i in range(1, 9)] == [
        "File Name",
        "File Type",
        "Description",
        "Date Found",
        "Last Seen",
        "Link",
        "Version",
        "Location",
    ]

    rows1 = _rows_by_location(ws1, h1["Location"])
    assert set(rows1.keys()) == {"a.pdf", "sub/b.xlsx", "sub/c.text"}

    r_a_1 = rows1["a.pdf"]
    assert ws1.cell(r_a_1, h1["File Name"]).value == "a.pdf"
    assert ws1.cell(r_a_1, h1["File Type"]).value == "PDF"
    assert (ws1.cell(r_a_1, h1["Description"]).value or "") == ""
    assert _cell_date(ws1, r_a_1, h1["Date Found"]) == d1
    assert _cell_date(ws1, r_a_1, h1["Last Seen"]) == d1
    assert ws1.cell(r_a_1, h1["Version"]).value == "1.0"

    link_cell_1 = ws1.cell(r_a_1, h1["Link"])
    assert link_cell_1.value == "a.pdf"
    assert link_cell_1.hyperlink is not None
    assert link_cell_1.hyperlink.target.endswith("/a.pdf")

    # Manually populate Description to ensure it is preserved
    ws1.cell(r_a_1, h1["Description"]).value = "Important doc"
    wb1.save(out)

    # --- 2) Update unchanged: preserve Date Found + Description; bump Last Seen
    d2 = date(2025, 12, 19)
    create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        sharepoint_base_url="https://sp.example/doclib",
        today=d2,
    )

    wb2 = load_workbook(out)
    ws2 = wb2[MAIN_SHEET_NAME]
    h2 = _headers(ws2)
    rows2 = _rows_by_location(ws2, h2["Location"])
    r_a_2 = rows2["a.pdf"]

    assert _cell_date(ws2, r_a_2, h2["Date Found"]) == d1
    assert _cell_date(ws2, r_a_2, h2["Last Seen"]) == d2
    assert ws2.cell(r_a_2, h2["Version"]).value == "1.0"
    assert ws2.cell(r_a_2, h2["Description"]).value == "Important doc"

    # Update with file content change: preserve Description, bump Version, update Last Seen, Date Found unchanged
    _write(root / "a.pdf", b"%PDF-1.4\nhello changed\n")
    d3 = date(2025, 12, 20)
    create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        sharepoint_base_url="https://sp.example/doclib",
        today=d3,
    )

    wb3 = load_workbook(out)
    ws3 = wb3[MAIN_SHEET_NAME]
    h3 = _headers(ws3)
    rows3 = _rows_by_location(ws3, h3["Location"])
    r_a_3 = rows3["a.pdf"]

    assert _cell_date(ws3, r_a_3, h3["Date Found"]) == d1
    assert _cell_date(ws3, r_a_3, h3["Last Seen"]) == d3
    assert ws3.cell(r_a_3, h3["Version"]).value == "1.1"
    assert ws3.cell(r_a_3, h3["Description"]).value == "Important doc"

    # --- 3) Missing file with prune_missing=False:
    # Row should remain and Last Seen should NOT advance for the missing file.
    (root / "sub" / "b.xlsx").unlink()

    d4 = date(2025, 12, 21)
    create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        sharepoint_base_url="https://sp.example/doclib",
        today=d4,
        prune_missing=False,
    )

    wb4 = load_workbook(out)
    ws4 = wb4[MAIN_SHEET_NAME]
    h4 = _headers(ws4)
    rows4 = _rows_by_location(ws4, h4["Location"])

    assert "sub/b.xlsx" in rows4  # still present in the sheet
    r_b_4 = rows4["sub/b.xlsx"]

    # Last Seen should remain whatever it was on the last run when it existed (d3)
    assert _cell_date(ws4, r_b_4, h4["Last Seen"]) == d3

    # Files that still exist should have Last Seen updated to d4
    assert _cell_date(ws4, rows4["a.pdf"], h4["Last Seen"]) == d4
    assert _cell_date(ws4, rows4["sub/c.text"], h4["Last Seen"]) == d4

    # --- 4) Deletion removed when prune_missing=True --------------------------
    d5 = date(2025, 12, 22)
    create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        sharepoint_base_url="https://sp.example/doclib",
        today=d5,
        prune_missing=True,
    )

    wb5 = load_workbook(out)
    ws5 = wb5[MAIN_SHEET_NAME]
    h5 = _headers(ws5)
    rows5 = _rows_by_location(ws5, h5["Location"])

    assert "sub/b.xlsx" not in rows5
    assert "a.pdf" in rows5
    assert "sub/c.text" in rows5

    # Remaining files should have Last Seen updated to d5
    assert _cell_date(ws5, rows5["a.pdf"], h5["Last Seen"]) == d5
    assert _cell_date(ws5, rows5["sub/c.text"], h5["Last Seen"]) == d5

    # Description should still be preserved
    assert ws5.cell(rows5["a.pdf"], h5["Description"]).value == "Important doc"


def test_treasureignore_directories_files_and_wildcards(tmp_path: Path) -> None:
    """
    .treasureignore entries for:
      - directories
      - individual files
      - wildcards
      - negation (!), last-match-wins
    """
    root = tmp_path / "root"
    root.mkdir()

    # Files we want to keep
    _write(root / "keep.pdf", b"keep")
    _write(root / "notes.txt", b"notes")
    _write(root / "sub" / "ok.docx", b"docx")

    # Files/dirs we will ignore via rules (use only suffixes that are in DEFAULT_DOCUMENT_SUFFIXES)
    _write(root / "secret.pdf", b"secret")  # individual file rule
    _write(root / "scratch.txt", b"scratch")  # wildcard *.txt rule (instead of *.tmp)
    _write(root / "build" / "out.pdf", b"build")  # directory rule build/
    _write(root / "sub" / "cache1" / "x.pdf", b"cache1")  # wildcard dir rule sub/cache*/
    _write(root / "sub" / "cache2" / "y.pdf", b"cache2")  # wildcard dir rule sub/cache*/
    _write(root / "sub" / "cache1" / "keep.pdf", b"override")  # negation example

    out = tmp_path / "treasure_map.xlsx"

    # First run WITHOUT ignore file: everything appears
    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    wb1 = load_workbook(out)
    ws1 = wb1[MAIN_SHEET_NAME]
    h1 = _headers(ws1)
    locs1 = set(_rows_by_location(ws1, h1["Location"]).keys())

    assert "keep.pdf" in locs1
    assert "notes.txt" in locs1
    assert "sub/ok.docx" in locs1
    assert "secret.pdf" in locs1
    assert "scratch.txt" in locs1
    assert "build/out.pdf" in locs1
    assert "sub/cache1/x.pdf" in locs1
    assert "sub/cache2/y.pdf" in locs1
    assert "sub/cache1/keep.pdf" in locs1

    # Add ignore rules:
    # - build/ directory
    # - secret.pdf individual file
    # - *.txt wildcard (will ignore notes.txt and scratch.txt)
    # - sub/cache*/ wildcard directories
    # - but allow one file back via negation
    (root / ".treasureignore").write_text(
        "\n".join(
            [
                "build/",
                "secret.pdf",
                "*.txt",
                "sub/cache*/",
                "!sub/cache1/keep.pdf",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    # Re-run: ignored entries should be removed from the map even without prune_missing
    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 19), prune_missing=False)

    wb2 = load_workbook(out)
    ws2 = wb2[MAIN_SHEET_NAME]
    h2 = _headers(ws2)
    locs2 = set(_rows_by_location(ws2, h2["Location"]).keys())

    # Kept
    assert "keep.pdf" in locs2
    assert "sub/ok.docx" in locs2

    # Ignored
    assert "secret.pdf" not in locs2
    assert "notes.txt" not in locs2
    assert "scratch.txt" not in locs2
    assert "build/out.pdf" not in locs2
    assert "sub/cache1/x.pdf" not in locs2
    assert "sub/cache2/y.pdf" not in locs2

    # Negation kept this one file under an ignored dir-pattern set
    assert "sub/cache1/keep.pdf" in locs2
