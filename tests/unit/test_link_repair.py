"""Link-repair tests.

Covers the user-reported bug: renaming the ROOT of a scanned tree left every
relative ``Location`` correct but every stored absolute ``file://`` hyperlink
pointing at the old root, so dof marked trivially-repairable rows ``Broken``.

Repair must be *targeted* -- only rows whose stored target fails to resolve and
whose file this scan actually found -- so a hand-edited but working hyperlink is
never clobbered, and a genuinely missing file is still reported broken.
"""

from __future__ import annotations

import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional

import pytest
from openpyxl import load_workbook

from dof import api
from dof.api import (
    STATUS_BROKEN,
    STATUS_MOVED,
    STATUS_OK,
    ScanResult,
    WriteOutcome,
    create_or_update_treasure_map,
)

DAY1 = date(2026, 1, 1)
DAY2 = date(2026, 1, 2)
DAY3 = date(2026, 1, 3)
SP_BASE = "https://example.sharepoint.com/sites/Team/Shared%20Documents"

# Mirrors the shape of the reported tree: several files across nested folders.
TREE = {
    "Approvals/x.pptx": b"XXX",
    "Approvals/y.docx": b"YYY",
    "Contracts/2025/deal.pdf": b"DEAL",
    "readme.docx": b"README",
}


@pytest.fixture(autouse=True)
def _ban_network(monkeypatch: pytest.MonkeyPatch) -> None:
    """Fail the test if anything in this module attempts a network call."""

    def _boom(*args: object, **kwargs: object) -> None:
        raise AssertionError("network access attempted during link repair")

    monkeypatch.setattr(urllib.request, "urlopen", _boom)


def _write(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def _build_tree(root: Path) -> None:
    for rel, data in TREE.items():
        _write(root / rel, data)


def _scan(root: Path, out: Path, *, day: date = DAY1, **kwargs: object) -> WriteOutcome:
    result = create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        today=day,
        with_result=True,
        **kwargs,  # type: ignore[arg-type]
    )
    assert isinstance(result, WriteOutcome)
    return result


def _rows(out: Path) -> Dict[str, Dict[str, object]]:
    """Return every row keyed by Location, including its hyperlink target."""
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = {c.value: c.column for c in ws[1] if c.value}
    rows: Dict[str, Dict[str, object]] = {}
    for r in range(2, ws.max_row + 1):
        loc = ws.cell(r, headers["Location"]).value
        if not loc:
            continue
        row: Dict[str, object] = {name: ws.cell(r, col).value for name, col in headers.items()}
        link_cell = ws.cell(r, headers["Link"])
        row["__target"] = link_cell.hyperlink.target if link_cell.hyperlink else None
        rows[str(loc)] = row
    return rows


def _edit_cell(out: Path, location: str, column: str, value: object) -> None:
    """Simulate a user edit to one cell of an existing workbook."""
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = {c.value: c.column for c in ws[1] if c.value}
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, headers["Location"]).value) == location:
            ws.cell(r, headers[column]).value = value
            wb.save(out)
            return
    raise AssertionError(f"no row at {location}")  # pragma: no cover


def _set_target(out: Path, location: str, target: Optional[str]) -> None:
    """Rewrite the stored hyperlink target of one row, as a hand edit would."""
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = {c.value: c.column for c in ws[1] if c.value}
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, headers["Location"]).value) == location:
            ws.cell(r, headers["Link"]).hyperlink = target
            wb.save(out)
            return
    raise AssertionError(f"no row at {location}")  # pragma: no cover


# ------------------------------------------------------- R1: the reported bug


def test_r1_renaming_the_root_repairs_every_link_instead_of_breaking_it(tmp_path: Path) -> None:
    """The reported scenario: rename the root, rescan, expect Repaired not Broken."""
    old_root = tmp_path / "A-TEAMS" / "Sales - EMEA - Prospects - CEVA"
    new_root = tmp_path / "A-TEAMS" / "CEVA" / "Sandalwood"
    out = tmp_path / "map.xlsx"  # deliberately outside the tree, so the rename spares it
    _build_tree(old_root)
    _scan(old_root, out, day=DAY1)

    before = _rows(out)
    assert set(before) == set(TREE)
    assert all(str(r["__target"]).startswith("file:") for r in before.values())
    _edit_cell(out, "readme.docx", "Description", "Kept by the user")

    new_root.parent.mkdir(parents=True, exist_ok=True)
    old_root.rename(new_root)
    outcome = _scan(new_root, out, day=DAY2)
    scan = outcome.scan

    # Every row repaired, none broken -> exit code stays 0 at the CLI layer.
    assert sorted(scan.repaired_links) == sorted(TREE)
    assert scan.broken_links == []
    assert [c.location for c in scan.changes if c.change_type is api.ChangeType.REPAIRED] == sorted(TREE, key=str.lower)
    assert "  Repaired:  4" in scan.summary()

    after = _rows(out)
    for loc in TREE:
        target = str(after[loc]["__target"])
        assert target == (new_root / loc).resolve().as_uri()
        assert "Sales - EMEA - Prospects - CEVA" not in urllib.parse.unquote(target)
        assert after[loc]["Status"] == STATUS_OK
        # History survives the repair.
        assert after[loc]["Date Found"] == before[loc]["Date Found"]
        assert after[loc]["Version"] == before[loc]["Version"] == "1.0"
        last_seen = after[loc]["Last Seen"]
        assert isinstance(last_seen, datetime)
        assert last_seen.date() == DAY2
    assert after["readme.docx"]["Description"] == "Kept by the user"


# ------------------------------------------------------- R2: SharePoint base


def test_r2_stale_sharepoint_target_is_regenerated_against_the_configured_base(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _build_tree(root)
    _scan(root, out, day=DAY1, sharepoint_base_url=SP_BASE)

    # A target under the configured base whose path no longer exists in the tree:
    # exactly what a renamed root leaves behind in a SharePoint-linked workbook.
    stale = SP_BASE + "/Sales%20-%20EMEA/Approvals/x.pptx"
    _set_target(out, "Approvals/x.pptx", stale)

    outcome = _scan(root, out, day=DAY2, sharepoint_base_url=SP_BASE)

    assert outcome.scan.repaired_links == ["Approvals/x.pptx"]
    assert outcome.scan.broken_links == []
    rows = _rows(out)
    assert rows["Approvals/x.pptx"]["__target"] == SP_BASE + "/Approvals/x.pptx"
    assert rows["Approvals/x.pptx"]["Status"] == STATUS_OK
    # Untouched rows keep the targets they already had.
    assert rows["readme.docx"]["__target"] == SP_BASE + "/readme.docx"


# ------------------------------------------------------- R3: targeted, not blanket


def test_r3_hand_edited_but_resolvable_link_is_left_untouched(tmp_path: Path) -> None:
    """Repair must never clobber a user's working hyperlink."""
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _build_tree(root)
    elsewhere = tmp_path / "elsewhere" / "canonical.docx"
    _write(elsewhere, b"CANON")
    _scan(root, out, day=DAY1)

    hand_edited = elsewhere.resolve().as_uri()
    _set_target(out, "readme.docx", hand_edited)

    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.repaired_links == []
    assert outcome.scan.broken_links == []
    rows = _rows(out)
    assert rows["readme.docx"]["__target"] == hand_edited
    assert rows["readme.docx"]["Status"] == STATUS_OK


def test_r3b_repair_preserves_the_displayed_link_text(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "readme.docx", b"README")
    _scan(root, out, day=DAY1)
    _edit_cell(out, "readme.docx", "Link", "Open the readme")
    _set_target(out, "readme.docx", (tmp_path / "gone" / "readme.docx").resolve().as_uri())

    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.repaired_links == ["readme.docx"]
    rows = _rows(out)
    assert rows["readme.docx"]["Link"] == "Open the readme"
    assert rows["readme.docx"]["__target"] == (root / "readme.docx").resolve().as_uri()


# ------------------------------------------------------- R4: still broken when unrepairable


def test_r4_missing_file_is_still_broken_and_never_repaired(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _build_tree(root)
    _scan(root, out, day=DAY1)

    (root / "readme.docx").unlink()
    outcome = _scan(root, out, day=DAY2, prune_missing=False)

    assert outcome.scan.repaired_links == []
    assert outcome.scan.broken_links == ["readme.docx"]
    assert _rows(out)["readme.docx"]["Status"] == STATUS_BROKEN
    assert "  Repaired:" not in outcome.scan.summary()


def test_r4b_root_rename_repairs_survivors_while_a_deleted_file_stays_broken(tmp_path: Path) -> None:
    """Mixed case: repair does not mask a genuinely unrepairable row."""
    old_root = tmp_path / "old"
    new_root = tmp_path / "new"
    out = tmp_path / "map.xlsx"
    _build_tree(old_root)
    _scan(old_root, out, day=DAY1)

    old_root.rename(new_root)
    (new_root / "readme.docx").unlink()
    outcome = _scan(new_root, out, day=DAY2, prune_missing=False)

    assert outcome.scan.broken_links == ["readme.docx"]
    assert sorted(outcome.scan.repaired_links) == sorted(set(TREE) - {"readme.docx"})


# ------------------------------------------------------- R5: Moved is not regressed


def test_r5_moved_status_is_not_regressed_by_the_repair_pass(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"AAA")
    _scan(root, out, day=DAY1)

    (root / "b").mkdir()
    (root / "a" / "doc.pdf").rename(root / "b" / "doc.pdf")
    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == [("a/doc.pdf", "b/doc.pdf")]
    assert outcome.scan.repaired_links == []
    assert outcome.scan.broken_links == []
    rows = _rows(out)
    assert rows["b/doc.pdf"]["Status"] == STATUS_MOVED
    assert rows["b/doc.pdf"]["__target"] == (root / "b" / "doc.pdf").resolve().as_uri()


def test_r5b_moved_row_with_a_stale_target_is_repaired_but_stays_moved(tmp_path: Path) -> None:
    """Repair rewrites the target; Status ownership stays with the move pass."""
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"AAA")
    _scan(root, out, day=DAY1)
    _set_target(out, "a/doc.pdf", (tmp_path / "nowhere" / "doc.pdf").resolve().as_uri())

    (root / "b").mkdir()
    (root / "a" / "doc.pdf").rename(root / "b" / "doc.pdf")
    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == [("a/doc.pdf", "b/doc.pdf")]
    assert outcome.scan.broken_links == []
    assert _rows(out)["b/doc.pdf"]["Status"] == STATUS_MOVED


# ------------------------------------------------------- R6: idempotence


def test_r6_a_second_scan_reports_no_repairs(tmp_path: Path) -> None:
    old_root = tmp_path / "old"
    new_root = tmp_path / "new"
    out = tmp_path / "map.xlsx"
    _build_tree(old_root)
    _scan(old_root, out, day=DAY1)

    old_root.rename(new_root)
    first = _scan(new_root, out, day=DAY2)
    assert sorted(first.scan.repaired_links) == sorted(TREE)

    second = _scan(new_root, out, day=DAY3)
    assert second.scan.repaired_links == []
    assert second.scan.broken_links == []
    assert [c for c in second.scan.changes if c.change_type is api.ChangeType.REPAIRED] == []
    assert "Repaired" not in second.scan.summary()


def test_r6b_dry_run_reports_repairs_without_writing(tmp_path: Path) -> None:
    old_root = tmp_path / "old"
    new_root = tmp_path / "new"
    out = tmp_path / "map.xlsx"
    _build_tree(old_root)
    _scan(old_root, out, day=DAY1)
    mtime = out.stat().st_mtime_ns

    old_root.rename(new_root)
    result = create_or_update_treasure_map(
        root_dir=new_root,
        output_xlsx=out,
        today=DAY2,
        dry_run=True,
    )

    assert isinstance(result, ScanResult)
    assert sorted(result.repaired_links) == sorted(TREE)
    assert result.broken_links == []
    assert out.stat().st_mtime_ns == mtime


# ------------------------------------------------------- R7: summary formatting


def test_r7_summary_omits_repaired_when_empty_and_appends_it_after_broken() -> None:
    plain = ScanResult(total_found=3, new_files=["a"], updated_files=["b"], unchanged_files=["c"])
    assert plain.summary() == ("Total documents found: 3\n  New:       1\n  Updated:   1\n  Unchanged: 1")
    assert "Repaired" not in plain.summary()

    repaired = ScanResult(
        total_found=3,
        new_files=["a"],
        updated_files=["b"],
        unchanged_files=["c"],
        broken_links=["gone"],
        repaired_links=["x", "y"],
    )
    lines = repaired.summary().splitlines()
    assert lines[:4] == plain.summary().splitlines()  # original lines byte-identical
    assert lines[-2:] == ["  Broken:    1", "  Repaired:  2"]


# ------------------------------------------------------- unit-level repair pass


def test_repair_link_targets_skips_rows_not_found_by_this_scan(tmp_path: Path) -> None:
    """Directly exercise the pass: only rows present in ``found_by_location`` are touched."""
    present = tmp_path / "here.pdf"
    present.write_bytes(b"x")
    found = api.FoundFile(
        abs_path=present,
        rel_location="here.pdf",
        filename="here.pdf",
        suffix=".pdf",
        file_type="PDF",
        sha256="deadbeef",
    )
    stale = (tmp_path / "old-root" / "here.pdf").resolve().as_uri()
    rows: Dict[str, Dict[str, object]] = {
        "here.pdf": {"Location": "here.pdf", "File Name": "here.pdf", "Link": "here.pdf", "__link_target": stale},
        "absent.pdf": {"Location": "absent.pdf", "File Name": "absent.pdf", "Link": "absent.pdf"},
    }
    scan = ScanResult()

    api._repair_link_targets(
        updated_rows=rows,
        found_by_location={"here.pdf": found},
        root_dir=tmp_path,
        found_locations={"here.pdf"},
        sharepoint_base_url=None,
        scan_result=scan,
    )

    assert scan.repaired_links == ["here.pdf"]
    assert rows["here.pdf"]["Link"] == {"target": present.resolve().as_uri(), "text": "here.pdf"}
    assert rows["absent.pdf"]["Link"] == "absent.pdf"  # untouched


def test_repair_link_targets_preserves_text_of_an_already_dict_link(tmp_path: Path) -> None:
    """A row relinked earlier in the run holds a dict Link; its text must survive repair."""
    present = tmp_path / "here.pdf"
    present.write_bytes(b"x")
    found = api.FoundFile(
        abs_path=present,
        rel_location="here.pdf",
        filename="here.pdf",
        suffix=".pdf",
        file_type="PDF",
        sha256="deadbeef",
    )
    stale = (tmp_path / "old-root" / "here.pdf").resolve().as_uri()
    rows: Dict[str, Dict[str, object]] = {
        "here.pdf": {
            "Location": "here.pdf",
            "File Name": "here.pdf",
            "Link": {"target": stale, "text": "Open me"},
        }
    }
    scan = ScanResult()

    api._repair_link_targets(
        updated_rows=rows,
        found_by_location={"here.pdf": found},
        root_dir=tmp_path,
        found_locations={"here.pdf"},
        sharepoint_base_url=None,
        scan_result=scan,
    )

    assert scan.repaired_links == ["here.pdf"]
    assert rows["here.pdf"]["Link"] == {"target": present.resolve().as_uri(), "text": "Open me"}
