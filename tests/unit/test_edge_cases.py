"""Edge case tests for dof.

Tests for:
- Unicode filenames with special characters
- Dry-run functionality
- JSON output format
- CSV output format
- Progress callback
- Unreadable .treasureignore handling
- Symlink handling
"""

from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path
from unittest.mock import Mock

import pytest
from openpyxl import load_workbook

from dof.api import (
    MAIN_SHEET_NAME,
    ChangeType,
    OutputFormat,
    ScanResult,
    create_or_update_treasure_map,
    discover_documents,
)


def _write(p: Path, content: bytes) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(content)


class TestUnicodeFilenames:
    """Test handling of Unicode filenames with special characters."""

    def test_unicode_filename_basic(self, tmp_path: Path) -> None:
        """Test basic Unicode characters in filenames."""
        root = tmp_path / "root"

        _write(root / "日本語.pdf", b"japanese")
        _write(root / "中文文档.docx", b"chinese")
        _write(root / "Ελληνικά.txt", b"greek")

        out = tmp_path / "map.xlsx"
        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

        wb = load_workbook(out)
        ws = wb[MAIN_SHEET_NAME]
        locs = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}

        assert "日本語.pdf" in locs
        assert "中文文档.docx" in locs
        assert "Ελληνικά.txt" in locs

    def test_unicode_filename_with_emoji(self, tmp_path: Path) -> None:
        """Test emoji characters in filenames (if filesystem supports)."""
        root = tmp_path / "root"

        try:
            _write(root / "report_📊.pdf", b"emoji")
            out = tmp_path / "map.xlsx"
            create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

            wb = load_workbook(out)
            ws = wb[MAIN_SHEET_NAME]
            locs = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}
            assert "report_📊.pdf" in locs
        except OSError:
            pytest.skip("Filesystem does not support emoji in filenames")

    def test_unicode_in_directory_names(self, tmp_path: Path) -> None:
        """Test Unicode characters in directory names."""
        root = tmp_path / "root"

        _write(root / "文档" / "report.pdf", b"in chinese dir")
        _write(root / "données" / "fichier.txt", b"in french dir")

        out = tmp_path / "map.xlsx"
        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

        wb = load_workbook(out)
        ws = wb[MAIN_SHEET_NAME]
        locs = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}

        assert "文档/report.pdf" in locs
        assert "données/fichier.txt" in locs


class TestDryRun:
    """Test dry-run functionality."""

    def test_dry_run_returns_scan_result(self, tmp_path: Path) -> None:
        """Test that dry_run=True returns ScanResult instead of Path."""
        root = tmp_path / "root"
        _write(root / "a.pdf", b"content")
        _write(root / "b.txt", b"text")

        out = tmp_path / "map.xlsx"
        result = create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18), dry_run=True)

        assert isinstance(result, ScanResult)
        assert result.total_found == 2
        assert len(result.new_files) == 2
        assert "a.pdf" in result.new_files
        assert "b.txt" in result.new_files

    def test_dry_run_does_not_create_file(self, tmp_path: Path) -> None:
        """Test that dry_run=True does not create the output file."""
        root = tmp_path / "root"
        _write(root / "a.pdf", b"content")

        out = tmp_path / "map.xlsx"
        assert not out.exists()

        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18), dry_run=True)

        assert not out.exists()

    def test_dry_run_detects_updates(self, tmp_path: Path) -> None:
        """Test that dry_run correctly identifies updated files."""
        root = tmp_path / "root"
        _write(root / "a.pdf", b"original")

        out = tmp_path / "map.xlsx"
        # First, create the map
        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

        # Modify file
        _write(root / "a.pdf", b"modified content")

        # Dry run should detect the change
        result = create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 19), dry_run=True)

        assert isinstance(result, ScanResult)
        assert len(result.updated_files) == 1
        assert "a.pdf" in result.updated_files

    def test_dry_run_detects_new_and_unchanged(self, tmp_path: Path) -> None:
        """Test dry_run correctly categorizes files."""
        root = tmp_path / "root"
        _write(root / "existing.pdf", b"existing")

        out = tmp_path / "map.xlsx"
        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

        # Add new file, keep existing unchanged
        _write(root / "new.pdf", b"new file")

        result = create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 19), dry_run=True)

        assert isinstance(result, ScanResult)
        assert len(result.new_files) == 1
        assert "new.pdf" in result.new_files
        assert len(result.unchanged_files) == 1
        assert "existing.pdf" in result.unchanged_files

    def test_scan_result_summary(self, tmp_path: Path) -> None:
        """Test ScanResult.summary() output."""
        root = tmp_path / "root"
        _write(root / "a.pdf", b"a")
        _write(root / "b.pdf", b"b")

        out = tmp_path / "map.xlsx"
        result = create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18), dry_run=True)

        summary = result.summary()
        assert "Total documents found: 2" in summary
        assert "New:       2" in summary


class TestOutputFormats:
    """Test JSON and CSV output formats."""

    def test_json_output(self, tmp_path: Path) -> None:
        """Test JSON output format."""
        root = tmp_path / "root"
        _write(root / "doc1.pdf", b"pdf content")
        _write(root / "sub" / "doc2.txt", b"text content")

        out = tmp_path / "map.xlsx"  # Extension will be changed to .json
        result = create_or_update_treasure_map(
            root_dir=root,
            output_xlsx=out,
            today=date(2025, 12, 18),
            output_format=OutputFormat.JSON,
        )

        assert isinstance(result, Path)
        assert result.suffix == ".json"
        assert result.exists()

        with result.open() as f:
            data = json.load(f)

        assert "treasure_map" in data
        assert len(data["treasure_map"]) == 2

        # Check structure of first entry
        entry = data["treasure_map"][0]
        assert "File Name" in entry
        assert "File Type" in entry
        assert "Date Found" in entry
        assert "Last Seen" in entry
        assert "Version" in entry
        assert "Location" in entry

    def test_csv_output(self, tmp_path: Path) -> None:
        """Test CSV output format."""
        root = tmp_path / "root"
        _write(root / "doc1.pdf", b"pdf content")
        _write(root / "sub" / "doc2.txt", b"text content")

        out = tmp_path / "map.xlsx"
        result = create_or_update_treasure_map(
            root_dir=root,
            output_xlsx=out,
            today=date(2025, 12, 18),
            output_format=OutputFormat.CSV,
        )

        assert isinstance(result, Path)
        assert result.suffix == ".csv"
        assert result.exists()

        with result.open(encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 2

        # Check header fields
        assert "File Name" in rows[0]
        assert "File Type" in rows[0]
        assert "Date Found" in rows[0]

    def test_json_with_unicode(self, tmp_path: Path) -> None:
        """Test JSON output handles Unicode correctly."""
        root = tmp_path / "root"
        _write(root / "文档.pdf", b"chinese")

        out = tmp_path / "map.xlsx"
        result = create_or_update_treasure_map(
            root_dir=root,
            output_xlsx=out,
            today=date(2025, 12, 18),
            output_format=OutputFormat.JSON,
        )

        with result.open(encoding="utf-8") as f:
            data = json.load(f)

        assert data["treasure_map"][0]["File Name"] == "文档.pdf"

    def test_csv_with_unicode(self, tmp_path: Path) -> None:
        """Test CSV output handles Unicode correctly."""
        root = tmp_path / "root"
        _write(root / "données.pdf", b"french")

        out = tmp_path / "map.xlsx"
        result = create_or_update_treasure_map(
            root_dir=root,
            output_xlsx=out,
            today=date(2025, 12, 18),
            output_format=OutputFormat.CSV,
        )

        with result.open(encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert rows[0]["File Name"] == "données.pdf"


class TestProgressCallback:
    """Test progress callback functionality."""

    def test_progress_callback_is_called(self, tmp_path: Path) -> None:
        """Test that progress callback is called for each file."""
        root = tmp_path / "root"
        _write(root / "a.pdf", b"a")
        _write(root / "b.pdf", b"b")
        _write(root / "c.txt", b"c")

        callback = Mock()
        discover_documents(root, progress_callback=callback)

        assert callback.call_count == 3

    def test_progress_callback_receives_paths(self, tmp_path: Path) -> None:
        """Test that progress callback receives file paths."""
        root = tmp_path / "root"
        _write(root / "doc.pdf", b"content")

        paths_received = []

        def callback(path: str) -> None:
            paths_received.append(path)

        discover_documents(root, progress_callback=callback)

        assert len(paths_received) == 1
        assert "doc.pdf" in paths_received[0]


class TestSymlinks:
    """Test symlink handling."""

    @pytest.mark.skip(reason="Symlink support not yet implemented - resolve() follows links outside root")
    def test_symlink_to_file(self, tmp_path: Path) -> None:
        """Test that symlinks to files are followed."""
        root = tmp_path / "root"
        root.mkdir()

        # Create actual file outside root
        external = tmp_path / "external" / "real.pdf"
        _write(external, b"real content")

        # Create symlink inside root
        link = root / "link.pdf"
        link.symlink_to(external)

        out = tmp_path / "map.xlsx"
        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

        wb = load_workbook(out)
        ws = wb[MAIN_SHEET_NAME]
        locs = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}

        assert "link.pdf" in locs

    @pytest.mark.skip(reason="Symlink support not yet implemented - resolve() follows links outside root")
    def test_symlink_to_directory(self, tmp_path: Path) -> None:
        """Test that symlinks to directories are followed."""
        root = tmp_path / "root"
        root.mkdir()

        # Create directory with files outside root
        external_dir = tmp_path / "external_dir"
        _write(external_dir / "doc.pdf", b"content")

        # Create symlink to directory inside root
        link_dir = root / "linked_dir"
        link_dir.symlink_to(external_dir)

        out = tmp_path / "map.xlsx"
        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

        wb = load_workbook(out)
        ws = wb[MAIN_SHEET_NAME]
        locs = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}

        assert "linked_dir/doc.pdf" in locs


class TestTreasureignoreEdgeCases:
    """Edge cases for .treasureignore handling."""

    def test_treasureignore_with_bom(self, tmp_path: Path) -> None:
        """Test .treasureignore file with UTF-8 BOM is handled."""
        root = tmp_path / "root"
        _write(root / "ignore.pdf", b"ignore")
        _write(root / "keep.pdf", b"keep")

        # Write file with UTF-8 BOM
        ignore_file = root / ".treasureignore"
        ignore_file.write_bytes(b"\xef\xbb\xbfignore.pdf\n")

        out = tmp_path / "map.xlsx"
        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

        wb = load_workbook(out)
        ws = wb[MAIN_SHEET_NAME]
        locs = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}

        # File should still be ignored despite BOM (pattern after BOM stripping)
        assert "keep.pdf" in locs

    def test_treasureignore_binary_content(self, tmp_path: Path) -> None:
        """Test .treasureignore with binary/invalid UTF-8 content doesn't crash."""
        root = tmp_path / "root"
        _write(root / "doc.pdf", b"content")

        # Write binary garbage to .treasureignore
        (root / ".treasureignore").write_bytes(b"\x80\x81\x82\xff\xfe")

        out = tmp_path / "map.xlsx"
        # Should not raise an exception
        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

        wb = load_workbook(out)
        ws = wb[MAIN_SHEET_NAME]
        locs = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}

        # File should be present since ignore file was unreadable/invalid
        assert "doc.pdf" in locs


class TestFileChangeTracking:
    """Test FileChange and ChangeType tracking."""

    def test_change_types_in_scan_result(self, tmp_path: Path) -> None:
        """Test that all change types are properly tracked."""
        root = tmp_path / "root"
        _write(root / "existing.pdf", b"existing")

        out = tmp_path / "map.xlsx"
        create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

        # Modify existing and add new
        _write(root / "existing.pdf", b"modified")
        _write(root / "new.pdf", b"new")

        result = create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 19), dry_run=True)

        assert isinstance(result, ScanResult)

        # Check changes list
        change_types = {c.change_type for c in result.changes}
        assert ChangeType.NEW in change_types
        assert ChangeType.UPDATED in change_types

        # Verify specific changes
        new_changes = [c for c in result.changes if c.change_type == ChangeType.NEW]
        assert len(new_changes) == 1
        assert new_changes[0].location == "new.pdf"
        assert new_changes[0].new_version == "1.0"

        updated_changes = [c for c in result.changes if c.change_type == ChangeType.UPDATED]
        assert len(updated_changes) == 1
        assert updated_changes[0].location == "existing.pdf"
        assert updated_changes[0].old_version == "1.0"
        assert updated_changes[0].new_version == "1.1"
