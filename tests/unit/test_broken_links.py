"""Broken-link detection tests (acceptance criteria B1-B7 of PLAN.md).

Link validation is entirely offline. This module bans network access outright so
that a regression which introduces a request fails loudly rather than silently
slowing the suite down.
"""

from __future__ import annotations

import urllib.request
from datetime import date
from pathlib import Path
from typing import Dict

import pytest
from openpyxl import load_workbook

from dof import api
from dof.api import (
    STATUS_BROKEN,
    STATUS_MOVED,
    STATUS_OK,
    ScanResult,
    WriteOutcome,
    create_or_update_treasure_map,
)

DAY1 = date(2026, 1, 1)
DAY2 = date(2026, 1, 2)
SP_BASE = "https://example.sharepoint.com/sites/Team/Shared%20Documents"


@pytest.fixture(autouse=True)
def _ban_network(monkeypatch: pytest.MonkeyPatch) -> None:
    """Fail the test if anything in this module attempts a network call."""

    def _boom(*args: object, **kwargs: object) -> None:
        raise AssertionError("network access attempted during link validation")

    monkeypatch.setattr(urllib.request, "urlopen", _boom)


def _write(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def _scan(root: Path, out: Path, *, day: date = DAY1, **kwargs: object) -> WriteOutcome:
    result = create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        today=day,
        with_result=True,
        **kwargs,  # type: ignore[arg-type]
    )
    assert isinstance(result, WriteOutcome)
    return result


def _rows(out: Path) -> Dict[str, Dict[str, object]]:
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = {c.value: c.column for c in ws[1] if c.value}
    rows: Dict[str, Dict[str, object]] = {}
    for r in range(2, ws.max_row + 1):
        loc = ws.cell(r, headers["Location"]).value
        if not loc:
            continue
        row: Dict[str, object] = {name: ws.cell(r, col).value for name, col in headers.items()}
        row["__row"] = r
        rows[str(loc)] = row
    return rows


def _fill_rgb(out: Path, location: str, column: str = "Location") -> object:
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = {c.value: c.column for c in ws[1] if c.value}
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, headers["Location"]).value) == location:
            return ws.cell(r, headers[column]).fill.start_color.rgb
    raise AssertionError(f"no row at {location}")  # pragma: no cover


# ---------------------------------------------------------------- B1 / B2


def test_b1_pruned_row_is_not_a_broken_link(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a.pdf", b"AAA")
    _write(root / "b.pdf", b"BBB")
    _scan(root, out, day=DAY1)

    (root / "b.pdf").unlink()
    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.broken_links == []
    assert outcome.scan.deleted_files == ["b.pdf"]
    assert list(_rows(out)) == ["a.pdf"]


def test_b2_keep_missing_marks_the_row_broken(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a.pdf", b"AAA")
    _write(root / "b.pdf", b"BBB")
    _scan(root, out, day=DAY1)

    (root / "b.pdf").unlink()
    outcome = _scan(root, out, day=DAY2, prune_missing=False)

    assert outcome.scan.broken_links == ["b.pdf"]
    rows = _rows(out)
    assert rows["b.pdf"]["Status"] == STATUS_BROKEN
    assert rows["a.pdf"]["Status"] == STATUS_OK
    assert any(c.change_type is api.ChangeType.BROKEN and c.location == "b.pdf" for c in outcome.scan.changes)


# ---------------------------------------------------------------- B3 / B4


def test_b3_broken_row_is_red_filled(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a.pdf", b"AAA")
    _write(root / "b.pdf", b"BBB")
    _scan(root, out, day=DAY1)

    (root / "b.pdf").unlink()
    _scan(root, out, day=DAY2, prune_missing=False)

    assert _fill_rgb(out, "b.pdf") == "FFFFC7CE"
    assert _fill_rgb(out, "b.pdf", column="File Name") == "FFFFC7CE"
    assert _fill_rgb(out, "a.pdf") in (None, "00000000")


def test_b4_reappearing_file_clears_broken_status_and_fill(tmp_path: Path) -> None:
    """The regression a naive implementation misses: stale red fill at a reused row index."""
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a.pdf", b"AAA")
    _write(root / "b.pdf", b"BBB")
    _scan(root, out, day=DAY1)

    (root / "b.pdf").unlink()
    broken = _scan(root, out, day=DAY2, prune_missing=False)
    assert broken.scan.broken_links == ["b.pdf"]
    broken_row_index = _rows(out)["b.pdf"]["__row"]

    _write(root / "b.pdf", b"BBB")
    healed = _scan(root, out, day=date(2026, 1, 3), prune_missing=False)

    assert healed.scan.broken_links == []
    rows = _rows(out)
    assert rows["b.pdf"]["Status"] == STATUS_OK
    assert rows["b.pdf"]["__row"] == broken_row_index  # same index -> stale formatting would persist
    assert _fill_rgb(out, "b.pdf") in (None, "00000000")


def test_b4b_moved_row_keeps_moved_status_when_resolvable(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"AAA")
    _scan(root, out, day=DAY1)
    (root / "b").mkdir()
    (root / "a" / "doc.pdf").rename(root / "b" / "doc.pdf")

    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.broken_links == []
    assert _rows(out)["b/doc.pdf"]["Status"] == STATUS_MOVED


# ---------------------------------------------------------------- B5 / B6


def test_b5_sharepoint_links_are_validated_offline(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a.pdf", b"AAA")
    _write(root / "b.pdf", b"BBB")
    _scan(root, out, day=DAY1, sharepoint_base_url=SP_BASE)

    (root / "b.pdf").unlink()
    outcome = _scan(root, out, day=DAY2, prune_missing=False, sharepoint_base_url=SP_BASE)

    assert outcome.scan.broken_links == ["b.pdf"]
    rows = _rows(out)
    assert rows["a.pdf"]["Status"] == STATUS_OK
    assert rows["b.pdf"]["Status"] == STATUS_BROKEN


def test_b6_hand_pasted_https_link_is_never_flagged_broken(tmp_path: Path) -> None:
    """dof cannot judge a URL it did not generate, so it must not call it broken."""
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a.pdf", b"AAA")
    _write(root / "b.pdf", b"BBB")
    _scan(root, out, day=DAY1, sharepoint_base_url="https://elsewhere.example.com/docs")

    (root / "b.pdf").unlink()
    # No base URL configured this run: the stored https target is user data.
    outcome = _scan(root, out, day=DAY2, prune_missing=False)

    assert outcome.scan.broken_links == []
    assert _rows(out)["b.pdf"]["Status"] == STATUS_OK


def test_link_is_resolvable_handles_file_uri_and_missing_location(tmp_path: Path) -> None:
    present = tmp_path / "present.pdf"
    present.write_bytes(b"x")
    missing = tmp_path / "gone.pdf"

    assert api._link_is_resolvable(
        {"Location": "present.pdf", "Link": {"target": present.as_uri()}},
        root_dir=tmp_path,
        found_locations={"present.pdf"},
        sharepoint_base_url=None,
    )
    assert not api._link_is_resolvable(
        {"Location": "gone.pdf", "Link": {"target": missing.as_uri()}},
        root_dir=tmp_path,
        found_locations=set(),
        sharepoint_base_url=None,
    )
    assert not api._link_is_resolvable(
        {"Location": "", "Link": ""},
        root_dir=tmp_path,
        found_locations=set(),
        sharepoint_base_url=None,
    )


# ---------------------------------------------------------------- B7


def test_b7_summary_lines() -> None:
    plain = ScanResult(total_found=3, new_files=["a"], updated_files=["b"], unchanged_files=["c"])
    assert plain.summary() == ("Total documents found: 3\n  New:       1\n  Updated:   1\n  Unchanged: 1")
    assert "Moved:" not in plain.summary()
    assert "Broken:" not in plain.summary()

    rich = ScanResult(
        total_found=3,
        new_files=["a"],
        updated_files=["b"],
        unchanged_files=["c"],
        moved_files=[("old", "new")],
        broken_links=["gone"],
    )
    lines = rich.summary().splitlines()
    assert lines[:4] == plain.summary().splitlines()[:4]
    assert "  Moved:     1" in lines
    assert "  Broken:    1" in lines


def test_dry_run_reports_broken_links_without_writing(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a.pdf", b"AAA")
    _write(root / "b.pdf", b"BBB")
    _scan(root, out, day=DAY1)
    mtime = out.stat().st_mtime_ns

    (root / "b.pdf").unlink()
    result = create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        today=DAY2,
        prune_missing=False,
        dry_run=True,
    )
    assert isinstance(result, ScanResult)
    assert result.broken_links == ["b.pdf"]
    assert out.stat().st_mtime_ns == mtime
