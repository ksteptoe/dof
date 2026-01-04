"""Comprehensive tests for .treasureignore pattern matching.

Tests gitignore-style patterns including:
- Root-anchored patterns
- Wildcard patterns (**, *, ?)
- Directory-only patterns
- Negation patterns (last-match-wins)
- Comments and blank lines
- Complex nested patterns
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from dof.api import MAIN_SHEET_NAME, create_or_update_treasure_map


def _write(p: Path, content: bytes) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(content)


def _get_locations(out: Path) -> set[str]:
    """Helper to extract all Location values from the treasure map."""
    wb = load_workbook(out)
    ws = wb[MAIN_SHEET_NAME]

    # Find Location column (should be column 8)
    loc_col = None
    for i in range(1, ws.max_column + 1):
        if ws.cell(1, i).value == "Location":
            loc_col = i
            break
    assert loc_col is not None, "Location column not found"

    return {str(ws.cell(r, loc_col).value) for r in range(2, ws.max_row + 1)}


def test_root_anchored_patterns(tmp_path: Path) -> None:
    """Test patterns starting with / that match only at root."""
    root = tmp_path / "root"

    _write(root / "exclude.pdf", b"root-exclude")
    _write(root / "keep.pdf", b"root-keep")
    _write(root / "sub" / "exclude.pdf", b"sub-exclude")  # same name, different location
    _write(root / "sub" / "keep.pdf", b"sub-keep")

    out = tmp_path / "map.xlsx"

    # Root-anchored pattern /exclude.pdf should only ignore at root
    (root / ".treasureignore").write_text("/exclude.pdf\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "exclude.pdf" not in locs  # ignored at root
    assert "keep.pdf" in locs
    assert "sub/exclude.pdf" in locs  # NOT ignored in subdirectory
    assert "sub/keep.pdf" in locs


def test_double_star_patterns(tmp_path: Path) -> None:
    """Test ** wildcard patterns that match across directory boundaries."""
    root = tmp_path / "root"

    _write(root / "a" / "temp" / "1.pdf", b"a-temp")
    _write(root / "a" / "b" / "temp" / "2.pdf", b"ab-temp")
    _write(root / "keep" / "3.pdf", b"keep")
    _write(root / "temp.pdf", b"temp-root")

    out = tmp_path / "map.xlsx"

    # Pattern **/temp/** should match any 'temp' directory at any depth
    (root / ".treasureignore").write_text("**/temp/**\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "a/temp/1.pdf" not in locs
    assert "a/b/temp/2.pdf" not in locs
    assert "keep/3.pdf" in locs
    assert "temp.pdf" in locs  # file named temp.pdf, not in temp/ directory


def test_basename_wildcard_patterns(tmp_path: Path) -> None:
    """Test patterns without / that match basename anywhere in tree."""
    root = tmp_path / "root"

    _write(root / "draft.pdf", b"draft-root")
    _write(root / "final.pdf", b"final-root")
    _write(root / "docs" / "draft.pdf", b"draft-docs")
    _write(root / "docs" / "final.pdf", b"final-docs")
    _write(root / "a" / "b" / "draft.pdf", b"draft-ab")

    out = tmp_path / "map.xlsx"

    # Pattern draft.pdf (no /) should match anywhere
    (root / ".treasureignore").write_text("draft.pdf\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "draft.pdf" not in locs
    assert "docs/draft.pdf" not in locs
    assert "a/b/draft.pdf" not in locs
    assert "final.pdf" in locs
    assert "docs/final.pdf" in locs


def test_directory_patterns_with_trailing_slash(tmp_path: Path) -> None:
    """Test directory-only patterns (trailing /) ignore entire directory trees."""
    root = tmp_path / "root"

    _write(root / "tmp" / "a.pdf", b"tmp-a")
    _write(root / "tmp" / "sub" / "b.pdf", b"tmp-sub-b")
    _write(root / "keep" / "tmp" / "c.pdf", b"keep-tmp-c")
    _write(root / "tmp.pdf", b"tmp-file")  # file named tmp.pdf
    _write(root / "docs" / "d.pdf", b"docs-d")

    out = tmp_path / "map.xlsx"

    # Pattern tmp/ should ignore all tmp directories (at any depth)
    (root / ".treasureignore").write_text("tmp/\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "tmp/a.pdf" not in locs
    assert "tmp/sub/b.pdf" not in locs
    assert "keep/tmp/c.pdf" not in locs
    assert "tmp.pdf" in locs  # file, not directory
    assert "docs/d.pdf" in locs


def test_negation_last_match_wins(tmp_path: Path) -> None:
    """Test negation patterns and last-match-wins semantics."""
    root = tmp_path / "root"

    _write(root / "docs" / "a.txt", b"a")
    _write(root / "docs" / "b.txt", b"b")
    _write(root / "docs" / "important.txt", b"important")
    _write(root / "docs" / "README.txt", b"readme")

    out = tmp_path / "map.xlsx"

    # Ignore all .txt but allow important.txt and README.txt back
    (root / ".treasureignore").write_text("*.txt\n!important.txt\n!README.txt\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "docs/a.txt" not in locs
    assert "docs/b.txt" not in locs
    assert "docs/important.txt" in locs
    assert "docs/README.txt" in locs


def test_negation_then_ignore_again(tmp_path: Path) -> None:
    """Test that last match wins: negation can be overridden by later ignore."""
    root = tmp_path / "root"

    _write(root / "docs" / "secret.pdf", b"secret")
    _write(root / "docs" / "public.pdf", b"public")

    out = tmp_path / "map.xlsx"

    # Allow secret.pdf, then ignore it again - last match wins
    (root / ".treasureignore").write_text("*.pdf\n!secret.pdf\nsecret.pdf\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "docs/secret.pdf" not in locs  # final ignore wins
    assert "docs/public.pdf" not in locs


def test_comments_and_blank_lines(tmp_path: Path) -> None:
    """Test that comments and blank lines are properly ignored."""
    root = tmp_path / "root"

    _write(root / "a.pdf", b"a")
    _write(root / "b.pdf", b"b")
    _write(root / "c.pdf", b"c")
    _write(root / "hash.pdf", b"hash")  # file named hash.pdf

    out = tmp_path / "map.xlsx"

    # File with comments and blank lines
    (root / ".treasureignore").write_text(
        "# This is a comment\n"
        "\n"
        "a.pdf\n"
        "  \n"  # blank line with spaces
        "# Another comment\n"
        "b.pdf\n"
        "\n"
        "# hash.pdf would be a comment, so actual file hash.pdf should NOT be ignored\n",
        encoding="utf-8",
    )

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "a.pdf" not in locs
    assert "b.pdf" not in locs
    assert "c.pdf" in locs
    assert "hash.pdf" in locs  # not ignored by comment


def test_complex_nested_patterns(tmp_path: Path) -> None:
    """Test complex patterns with multiple wildcards and path segments."""
    root = tmp_path / "root"

    _write(root / "src" / "temp" / "file.pdf", b"src-temp")
    _write(root / "src" / "cache" / "file.pdf", b"src-cache")
    _write(root / "src" / "main" / "file.pdf", b"src-main")
    _write(root / "test" / "temp" / "file.pdf", b"test-temp")
    _write(root / "docs" / "file.pdf", b"docs")

    out = tmp_path / "map.xlsx"

    # Ignore src/temp and src/cache directories specifically
    (root / ".treasureignore").write_text("src/temp/\nsrc/cache/\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "src/temp/file.pdf" not in locs
    assert "src/cache/file.pdf" not in locs
    assert "src/main/file.pdf" in locs
    assert "test/temp/file.pdf" in locs  # different parent path
    assert "docs/file.pdf" in locs


def test_wildcard_in_directory_name(tmp_path: Path) -> None:
    """Test wildcards within directory names like cache*."""
    root = tmp_path / "root"

    _write(root / "cache1" / "a.pdf", b"cache1")
    _write(root / "cache2" / "b.pdf", b"cache2")
    _write(root / "cache_old" / "c.pdf", b"cache_old")
    _write(root / "mycache" / "d.pdf", b"mycache")
    _write(root / "data" / "e.pdf", b"data")

    out = tmp_path / "map.xlsx"

    # Pattern cache* should match directories starting with cache
    (root / ".treasureignore").write_text("cache*/\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "cache1/a.pdf" not in locs
    assert "cache2/b.pdf" not in locs
    assert "cache_old/c.pdf" not in locs
    assert "mycache/d.pdf" in locs  # doesn't start with cache
    assert "data/e.pdf" in locs


def test_pattern_with_multiple_extensions(tmp_path: Path) -> None:
    """Test patterns matching multiple file extensions."""
    root = tmp_path / "root"

    _write(root / "doc1.pdf", b"pdf1")
    _write(root / "doc2.txt", b"txt1")
    _write(root / "doc3.docx", b"docx1")
    _write(root / "doc4.xlsx", b"xlsx1")

    out = tmp_path / "map.xlsx"

    # Ignore specific extensions
    (root / ".treasureignore").write_text("*.txt\n*.docx\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "doc1.pdf" in locs
    assert "doc2.txt" not in locs
    assert "doc3.docx" not in locs
    assert "doc4.xlsx" in locs


def test_subdirectory_specific_pattern(tmp_path: Path) -> None:
    """Test patterns that match specific subdirectories only."""
    root = tmp_path / "root"

    _write(root / "docs" / "draft" / "a.pdf", b"docs-draft-a")
    _write(root / "docs" / "final" / "b.pdf", b"docs-final-b")
    _write(root / "projects" / "draft" / "c.pdf", b"proj-draft-c")
    _write(root / "draft" / "d.pdf", b"root-draft-d")

    out = tmp_path / "map.xlsx"

    # Only ignore docs/draft/ specifically, not all draft/ directories
    (root / ".treasureignore").write_text("docs/draft/\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "docs/draft/a.pdf" not in locs
    assert "docs/final/b.pdf" in locs
    assert "projects/draft/c.pdf" in locs  # different parent
    assert "draft/d.pdf" in locs  # at root


def test_empty_treasureignore_file(tmp_path: Path) -> None:
    """Test that empty .treasureignore file doesn't break anything."""
    root = tmp_path / "root"

    _write(root / "a.pdf", b"a")
    _write(root / "b.pdf", b"b")

    out = tmp_path / "map.xlsx"

    (root / ".treasureignore").write_text("", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "a.pdf" in locs
    assert "b.pdf" in locs


def test_only_comments_treasureignore(tmp_path: Path) -> None:
    """Test .treasureignore with only comments and blank lines."""
    root = tmp_path / "root"

    _write(root / "a.pdf", b"a")
    _write(root / "b.pdf", b"b")

    out = tmp_path / "map.xlsx"

    (root / ".treasureignore").write_text("# Just comments\n\n# And blank lines\n  \n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "a.pdf" in locs
    assert "b.pdf" in locs


def test_whitespace_around_patterns(tmp_path: Path) -> None:
    """Test that patterns with leading/trailing whitespace are properly trimmed."""
    root = tmp_path / "root"

    _write(root / "ignore.pdf", b"ignore")
    _write(root / "keep.pdf", b"keep")

    out = tmp_path / "map.xlsx"

    # Pattern with spaces around it
    (root / ".treasureignore").write_text("  ignore.pdf  \n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "ignore.pdf" not in locs
    assert "keep.pdf" in locs


def test_case_sensitivity(tmp_path: Path) -> None:
    """Test that pattern matching respects filesystem case sensitivity."""
    root = tmp_path / "root"

    _write(root / "Test.pdf", b"test-upper")
    _write(root / "test.pdf", b"test-lower")
    _write(root / "TEST.pdf", b"test-all-upper")

    out = tmp_path / "map.xlsx"

    # Pattern test.pdf should match exactly (case-sensitive on case-sensitive FS)
    (root / ".treasureignore").write_text("test.pdf\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    # On case-sensitive filesystems, only test.pdf would be ignored
    # On case-insensitive (Windows), all variants might be ignored
    # Let's just check that at least test.pdf is not in the results
    assert "test.pdf" not in locs


def test_deeply_nested_directory_pattern(tmp_path: Path) -> None:
    """Test pattern matching in deeply nested directory structures."""
    root = tmp_path / "root"

    _write(root / "a" / "b" / "c" / "d" / "e" / "deep.pdf", b"deep")
    _write(root / "a" / "b" / "c" / "d" / "shallow.pdf", b"shallow")
    _write(root / "x" / "y" / "z" / "other.pdf", b"other")

    out = tmp_path / "map.xlsx"

    # Pattern matching deep nesting
    (root / ".treasureignore").write_text("a/b/c/d/e/\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "a/b/c/d/e/deep.pdf" not in locs
    assert "a/b/c/d/shallow.pdf" in locs
    assert "x/y/z/other.pdf" in locs


def test_question_mark_wildcard(tmp_path: Path) -> None:
    """Test single-character wildcard ? in patterns."""
    root = tmp_path / "root"

    _write(root / "test1.pdf", b"test1")
    _write(root / "test2.pdf", b"test2")
    _write(root / "test10.pdf", b"test10")
    _write(root / "testa.pdf", b"testa")
    _write(root / "keep.pdf", b"keep")

    out = tmp_path / "map.xlsx"

    # Pattern test?.pdf should match test1.pdf, test2.pdf, testa.pdf but not test10.pdf
    (root / ".treasureignore").write_text("test?.pdf\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "test1.pdf" not in locs
    assert "test2.pdf" not in locs
    assert "testa.pdf" not in locs
    assert "test10.pdf" in locs  # two characters, not one
    assert "keep.pdf" in locs


def test_negation_with_whitespace(tmp_path: Path) -> None:
    """Test that negation patterns handle whitespace correctly."""
    root = tmp_path / "root"

    _write(root / "a.pdf", b"a")
    _write(root / "b.pdf", b"b")

    out = tmp_path / "map.xlsx"

    # Negation pattern with whitespace
    (root / ".treasureignore").write_text(
        "*.pdf\n  !b.pdf  \n",  # spaces around negation pattern
        encoding="utf-8",
    )

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))
    locs = _get_locations(out)

    assert "a.pdf" not in locs
    assert "b.pdf" in locs  # negation should work despite whitespace
