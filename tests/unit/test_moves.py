"""Move-detection tests (acceptance criteria M1-M12 and BC2 of PLAN.md).

The three tiers under test are:

* Tier 1 -- identical SHA-256 relinks the row and keeps its ``Version``.
* Tier 2 -- identical file name, byte size and file type relinks the row and
  bumps its ``Version``.
* Tier 3 -- anything left over becomes a brand-new row.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional

import pytest
from openpyxl import load_workbook

from dof import api
from dof.api import (
    STATUS_BROKEN,
    STATUS_MOVED,
    STATUS_OK,
    OutputFormat,
    WriteOutcome,
    create_or_update_treasure_map,
)

DAY1 = date(2026, 1, 1)
DAY2 = date(2026, 1, 2)


def _scan(root: Path, out: Path, *, day: date = DAY1, **kwargs: object) -> WriteOutcome:
    """Run a real (non-dry) scan and return the :class:`WriteOutcome`."""
    result = create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        today=day,
        with_result=True,
        **kwargs,  # type: ignore[arg-type]
    )
    assert isinstance(result, WriteOutcome)
    return result


def _rows(out: Path) -> Dict[str, Dict[str, object]]:
    """Read the saved workbook back as ``{Location: {column: value}}``."""
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = {c.value: c.column for c in ws[1] if c.value}
    rows: Dict[str, Dict[str, object]] = {}
    for r in range(2, ws.max_row + 1):
        loc = ws.cell(r, headers["Location"]).value
        if not loc:
            continue
        row: Dict[str, object] = {name: ws.cell(r, col).value for name, col in headers.items()}
        row["__row"] = r
        rows[str(loc)] = row
    return rows


def _as_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _set_description(out: Path, location: str, text: str) -> None:
    """Simulate a user typing into the Description column."""
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = {c.value: c.column for c in ws[1] if c.value}
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, headers["Location"]).value) == location:
            ws.cell(r, headers["Description"], text)
            break
    else:  # pragma: no cover - defensive
        raise AssertionError(f"no row at {location}")
    wb.save(out)


def _write(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


# ---------------------------------------------------------------- M1


def test_m1_pure_move_preserves_history(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"%PDF-1.4 content A\n")

    _scan(root, out, day=DAY1)
    _set_description(out, "a/doc.pdf", "mine")

    (root / "b").mkdir()
    (root / "a" / "doc.pdf").rename(root / "b" / "doc.pdf")

    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == [("a/doc.pdf", "b/doc.pdf")]
    assert outcome.scan.new_files == []
    assert outcome.scan.deleted_files == []

    rows = _rows(out)
    assert list(rows) == ["b/doc.pdf"]
    row = rows["b/doc.pdf"]
    assert _as_date(row["Date Found"]) == DAY1
    assert _as_date(row["Last Seen"]) == DAY2
    assert row["Description"] == "mine"
    assert row["Version"] == "1.0"
    assert row["Status"] == STATUS_MOVED
    assert row["Previous Location"] == "a/doc.pdf"


def test_m1b_previous_location_is_sticky_after_a_clean_scan(tmp_path: Path) -> None:
    """``Previous Location`` is historical: a later clean scan must not clear it."""
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")
    _scan(root, out, day=DAY1)
    (root / "b").mkdir()
    (root / "a" / "doc.pdf").rename(root / "b" / "doc.pdf")
    _scan(root, out, day=DAY2)

    outcome = _scan(root, out, day=date(2026, 1, 3))
    assert outcome.scan.moved_files == []

    row = _rows(out)["b/doc.pdf"]
    assert row["Status"] == STATUS_OK
    assert row["Previous Location"] == "a/doc.pdf"


# ---------------------------------------------------------------- M2


def test_m2_rename_in_place(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "old.pdf", b"same bytes")
    _scan(root, out, day=DAY1)

    (root / "a" / "old.pdf").rename(root / "a" / "new.pdf")
    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == [("a/old.pdf", "a/new.pdf")]
    row = _rows(out)["a/new.pdf"]
    assert row["File Name"] == "new.pdf"
    assert row["Version"] == "1.0"
    assert row["Previous Location"] == "a/old.pdf"


# ---------------------------------------------------------------- M3 / M4


def test_m3_move_with_edit_same_size_is_tier2_and_bumps_version(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.txt", b"AAAAAAAA")
    _scan(root, out, day=DAY1)

    (root / "a" / "doc.txt").unlink()
    _write(root / "b" / "doc.txt", b"BBBBBBBB")  # same length, different content

    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == [("a/doc.txt", "b/doc.txt")]
    row = _rows(out)["b/doc.txt"]
    assert row["Version"] == "1.1"
    assert row["Status"] == STATUS_MOVED
    assert row["Previous Location"] == "a/doc.txt"


def test_m4_move_with_edit_different_size_is_not_a_move(tmp_path: Path) -> None:
    """Documents the honest limit of Tier 2: different size means no evidence."""
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.txt", b"AAAAAAAA")
    _scan(root, out, day=DAY1)

    (root / "a" / "doc.txt").unlink()
    _write(root / "b" / "doc.txt", b"BBBBBBBBBBBBBBBB")

    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == []
    assert outcome.scan.deleted_files == ["a/doc.txt"]
    assert outcome.scan.new_files == ["b/doc.txt"]
    assert list(_rows(out)) == ["b/doc.txt"]


# ---------------------------------------------------------------- M5


def test_m5_swap_of_two_paths_keeps_both_rows(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "x.pdf", b"content-X")
    _write(root / "b" / "y.pdf", b"content-Y")
    _scan(root, out, day=DAY1)
    _set_description(out, "a/x.pdf", "desc-X")
    _set_description(out, "b/y.pdf", "desc-Y")

    # x moves into b/, y moves into a/ -- the two files swap directories.
    (root / "a" / "x.pdf").rename(root / "b" / "x.pdf")
    (root / "b" / "y.pdf").rename(root / "a" / "y.pdf")

    outcome = _scan(root, out, day=DAY2)

    assert sorted(outcome.scan.moved_files) == [
        ("a/x.pdf", "b/x.pdf"),
        ("b/y.pdf", "a/y.pdf"),
    ]
    rows = _rows(out)
    assert sorted(rows) == ["a/y.pdf", "b/x.pdf"]
    # Descriptions follow the content, not the path.
    assert rows["b/x.pdf"]["Description"] == "desc-X"
    assert rows["a/y.pdf"]["Description"] == "desc-Y"
    assert outcome.scan.deleted_files == []
    assert outcome.scan.new_files == []


def test_m5b_content_swapped_between_two_existing_paths_is_two_updates(tmp_path: Path) -> None:
    """Both paths still exist, so the same-location pass wins: two updates, no moves."""
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a.pdf", b"content-A")
    _write(root / "b.pdf", b"content-BB")
    _scan(root, out, day=DAY1)

    (root / "a.pdf").write_bytes(b"content-BB")
    (root / "b.pdf").write_bytes(b"content-A")

    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == []
    assert sorted(outcome.scan.updated_files) == ["a.pdf", "b.pdf"]
    rows = _rows(out)
    assert sorted(rows) == ["a.pdf", "b.pdf"]
    assert rows["a.pdf"]["Version"] == "1.1"


# ---------------------------------------------------------------- M6 / M7 / M8


def test_m6_duplicate_content_rows_both_move_without_merging(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "one.pdf", b"identical bytes")
    _write(root / "a" / "two.pdf", b"identical bytes")
    _scan(root, out, day=DAY1)

    (root / "c").mkdir()
    (root / "a" / "one.pdf").rename(root / "c" / "one.pdf")
    (root / "a" / "two.pdf").rename(root / "c" / "two.pdf")

    outcome = _scan(root, out, day=DAY2)

    rows = _rows(out)
    assert sorted(rows) == ["c/one.pdf", "c/two.pdf"]
    assert len(rows) == 2
    assert sorted(outcome.scan.moved_files) == [
        ("a/one.pdf", "c/one.pdf"),
        ("a/two.pdf", "c/two.pdf"),
    ]
    # No FoundFile consumed twice, no row merged away.
    new_locations = [new for _old, new in outcome.scan.moved_files]
    assert len(set(new_locations)) == len(new_locations)


def test_m7_ambiguous_hash_uses_filename_tie_break(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "x" / "report.pdf", b"same content everywhere")
    _write(root / "x" / "copy.pdf", b"same content everywhere")
    _scan(root, out, day=DAY1)
    _set_description(out, "x/report.pdf", "the report")

    (root / "y").mkdir()
    (root / "x" / "report.pdf").rename(root / "y" / "report.pdf")
    (root / "x" / "copy.pdf").rename(root / "y" / "copy.pdf")

    outcome = _scan(root, out, day=DAY2)

    assert ("x/report.pdf", "y/report.pdf") in outcome.scan.moved_files
    assert ("x/copy.pdf", "y/copy.pdf") in outcome.scan.moved_files
    assert _rows(out)["y/report.pdf"]["Description"] == "the report"


def test_m8_one_file_copied_into_two_locations(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"copied content")
    _scan(root, out, day=DAY1)

    (root / "a" / "doc.pdf").unlink()
    _write(root / "b" / "doc.pdf", b"copied content")
    _write(root / "c" / "doc.pdf", b"copied content")

    outcome = _scan(root, out, day=DAY2)

    # Deterministic: the lowest sorted found path wins the move.
    assert outcome.scan.moved_files == [("a/doc.pdf", "b/doc.pdf")]
    assert outcome.scan.new_files == ["c/doc.pdf"]
    rows = _rows(out)
    assert sorted(rows) == ["b/doc.pdf", "c/doc.pdf"]
    assert rows["b/doc.pdf"]["Status"] == STATUS_MOVED
    assert rows["c/doc.pdf"]["Status"] == STATUS_OK
    assert rows["c/doc.pdf"]["Previous Location"] in ("", None)


def test_m8b_pairing_is_deterministic_across_identical_runs(tmp_path: Path) -> None:
    """Two independent, identical trees must produce byte-identical move lists."""
    results = []
    for run in ("first", "second"):
        root = tmp_path / run / "root"
        out = tmp_path / run / "map.xlsx"
        _write(root / "a" / "doc.pdf", b"copied content")
        _write(root / "a" / "other.pdf", b"copied content")
        _scan(root, out, day=DAY1)
        (root / "a" / "doc.pdf").unlink()
        (root / "a" / "other.pdf").unlink()
        _write(root / "z" / "doc.pdf", b"copied content")
        _write(root / "z" / "other.pdf", b"copied content")
        results.append(_scan(root, out, day=DAY2).scan.moved_files)

    assert results[0] == results[1]


# ---------------------------------------------------------------- M9


def test_m9_move_into_ignored_directory_is_not_a_move(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"archived content")
    (root / ".treasureignore").write_text("archive/\n", encoding="utf-8")
    _scan(root, out, day=DAY1)

    (root / "archive").mkdir()
    (root / "a" / "doc.pdf").rename(root / "archive" / "doc.pdf")

    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == []
    assert outcome.scan.deleted_files == ["a/doc.pdf"]
    assert _rows(out) == {}


# ---------------------------------------------------------------- M10


def test_m10_unhashable_file_falls_through_to_tier2(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.txt", b"AAAAAAAA")
    _scan(root, out, day=DAY1)

    (root / "b").mkdir()
    (root / "a" / "doc.txt").rename(root / "b" / "doc.txt")

    def _no_hash(path: Path) -> Optional[str]:
        return None

    monkeypatch.setattr(api, "_safe_sha256_file", _no_hash)
    outcome = _scan(root, out, day=DAY2)

    # No Tier 1 evidence (hash is None); Tier 2 still pairs on name+size+type.
    assert outcome.scan.moved_files == [("a/doc.txt", "b/doc.txt")]
    assert _rows(out)["b/doc.txt"]["Version"] == "1.1"


def test_m10b_unhashable_file_with_different_name_does_not_mispair(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.txt", b"AAAAAAAA")
    _scan(root, out, day=DAY1)

    (root / "a" / "doc.txt").unlink()
    _write(root / "b" / "unrelated.txt", b"BBBBBBBB")  # same size, different name

    monkeypatch.setattr(api, "_safe_sha256_file", lambda path: None)
    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == []
    assert outcome.scan.new_files == ["b/unrelated.txt"]
    assert outcome.scan.deleted_files == ["a/doc.txt"]


# ---------------------------------------------------------------- M11


def test_m11_zero_byte_files_never_pair_in_either_tier(tmp_path: Path) -> None:
    """Every empty file shares one SHA-256 and one size, so neither tier may pair them."""
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "empty.txt", b"")
    _write(root / "keep.txt", b"kept")
    _scan(root, out, day=DAY1)

    (root / "a" / "empty.txt").unlink()
    _write(root / "b" / "empty.txt", b"")  # identical name, identical (zero) size

    outcome = _scan(root, out, day=DAY2)

    assert outcome.scan.moved_files == []
    assert outcome.scan.deleted_files == ["a/empty.txt"]
    assert outcome.scan.new_files == ["b/empty.txt"]
    assert _rows(out)["b/empty.txt"]["Version"] == "1.0"


# ---------------------------------------------------------------- M12


def test_m12_detect_moves_false_reproduces_old_behaviour(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")
    _scan(root, out, day=DAY1)
    _set_description(out, "a/doc.pdf", "mine")

    (root / "b").mkdir()
    (root / "a" / "doc.pdf").rename(root / "b" / "doc.pdf")

    outcome = _scan(root, out, day=DAY2, detect_moves=False)

    assert outcome.scan.moved_files == []
    assert outcome.scan.new_files == ["b/doc.pdf"]
    assert outcome.scan.deleted_files == ["a/doc.pdf"]
    rows = _rows(out)
    assert list(rows) == ["b/doc.pdf"]
    assert rows["b/doc.pdf"]["Description"] in ("", None)
    assert rows["b/doc.pdf"]["Status"] == STATUS_OK
    assert _as_date(rows["b/doc.pdf"]["Date Found"]) == DAY2


# ---------------------------------------------------------------- BC2


def test_bc2_json_export_carries_new_columns_in_order(tmp_path: Path) -> None:
    import json

    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")
    outcome = _scan(root, out, day=DAY1, output_format=OutputFormat.JSON)

    data = json.loads(Path(outcome.path).read_text(encoding="utf-8"))
    entry = data["treasure_map"][0]
    assert list(entry) == api.REQUIRED_COLUMNS
    assert entry["Status"] == STATUS_OK
    assert entry["Previous Location"] == ""


def test_bc2b_csv_export_carries_new_columns_in_order(tmp_path: Path) -> None:
    import csv

    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")
    outcome = _scan(root, out, day=DAY1, output_format=OutputFormat.CSV)

    with Path(outcome.path).open(encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        first = next(reader)
    assert header == api.REQUIRED_COLUMNS
    assert first[header.index("Status")] == STATUS_OK


# ---------------------------------------------------------------- meta sheet


def test_meta_sheet_records_size(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "map.xlsx"
    _write(root / "doc.txt", b"12345")
    _scan(root, out, day=DAY1)

    wb = load_workbook(out)
    meta_ws = wb[api.META_SHEET_NAME]
    assert [c.value for c in meta_ws[1]][:3] == ["Location", "Sha256", "Size"]
    assert meta_ws.cell(2, 1).value == "doc.txt"
    assert meta_ws.cell(2, 3).value == 5


@pytest.mark.parametrize("status", [STATUS_OK, STATUS_MOVED, STATUS_BROKEN])
def test_status_constants_are_distinct(status: str) -> None:
    assert isinstance(status, str) and status
    assert len({STATUS_OK, STATUS_MOVED, STATUS_BROKEN}) == 3
