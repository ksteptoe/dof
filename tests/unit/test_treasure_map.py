from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from dof.api import MAIN_SHEET_NAME, create_or_update_treasure_map


def _write(p: Path, content: bytes) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(content)


def test_create_and_update_treasure_map(tmp_path: Path) -> None:
    root = tmp_path / "root"
    _write(root / "a.pdf", b"%PDF-1.4\nhello\n")
    _write(root / "sub" / "b.xlsx", b"fake-xlsx-bytes")
    _write(root / "sub" / "c.text", b"hi")

    out = tmp_path / "treasure_map.xlsx"
    d1 = date(2025, 12, 18)
    create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        sharepoint_base_url="https://sp.example/doclib",
        today=d1,
    )

    wb = load_workbook(out)
    ws = wb[MAIN_SHEET_NAME]
    headers = [c.value for c in ws[1][:8]]
    assert headers == [
        "File Name",
        "File Type",
        "Description",
        "Date Found",
        "Last Seen",
        "Link",
        "Version",
        "Location",
    ]

    # 3 rows + header
    assert ws.max_row == 4

    # Find row by Location (Location is now column 8)
    rows = {ws.cell(r, 8).value: r for r in range(2, ws.max_row + 1)}
    assert "a.pdf" in rows
    assert "sub/b.xlsx" in rows
    assert "sub/c.text" in rows

    r = rows["a.pdf"]
    assert ws.cell(r, 1).value == "a.pdf"
    assert ws.cell(r, 2).value == "PDF"
    assert (ws.cell(r, 3).value or "") == ""

    df_cell = ws.cell(r, 4).value
    ls_cell = ws.cell(r, 5).value
    assert df_cell is not None
    assert ls_cell is not None
    assert df_cell.date() == d1  # Date Found (first seen)
    assert ls_cell.date() == d1  # Last Seen

    assert ws.cell(r, 7).value == "1.0"  # Version

    link_cell = ws.cell(r, 6)  # Link is now column 6
    assert link_cell.value == "a.pdf"
    assert link_cell.hyperlink is not None
    assert link_cell.hyperlink.target.endswith("/a.pdf")

    # Run again unchanged with a different date:
    # Date Found stays the same, Version stays the same, Last Seen updates
    d2 = date(2025, 12, 19)
    create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        sharepoint_base_url="https://sp.example/doclib",
        today=d2,
    )
    wb2 = load_workbook(out)
    ws2 = wb2[MAIN_SHEET_NAME]
    rows2 = {ws2.cell(r, 8).value: r for r in range(2, ws2.max_row + 1)}
    r2 = rows2["a.pdf"]

    df2 = ws2.cell(r2, 4).value
    ls2 = ws2.cell(r2, 5).value
    assert df2 is not None
    assert ls2 is not None
    assert df2.date() == d1  # Date Found unchanged
    assert ls2.date() == d2  # Last Seen updated
    assert ws2.cell(r2, 7).value == "1.0"  # Version unchanged

    # Modify file and re-run:
    # Date Found stays first-seen, Version bumps, Last Seen updates
    _write(root / "a.pdf", b"%PDF-1.4\nhello changed\n")
    d3 = date(2025, 12, 20)
    create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        sharepoint_base_url="https://sp.example/doclib",
        today=d3,
    )

    wb3 = load_workbook(out)
    ws3 = wb3[MAIN_SHEET_NAME]
    rows3 = {ws3.cell(r, 8).value: r for r in range(2, ws3.max_row + 1)}
    r3 = rows3["a.pdf"]

    df3 = ws3.cell(r3, 4).value
    ls3 = ws3.cell(r3, 5).value
    assert df3 is not None
    assert ls3 is not None
    assert df3.date() == d1  # Date Found unchanged
    assert ls3.date() == d3  # Last Seen updated
    assert ws3.cell(r3, 7).value == "1.1"  # Version bumped

    # Modify file and re-run:
    # Date Found stays first-seen, Version bumps, Last Seen updates
    _write(root / "a.pdf", b"%PDF-1.4\nhello changed\n")
    d3 = date(2025, 12, 20)
    create_or_update_treasure_map(
        root_dir=root,
        output_xlsx=out,
        sharepoint_base_url="https://sp.example/doclib",
        today=d3,
    )

    wb3 = load_workbook(out)
    ws3 = wb3[MAIN_SHEET_NAME]
    rows3 = {ws3.cell(r, 8).value: r for r in range(2, ws3.max_row + 1)}
    r3 = rows3["a.pdf"]
    assert ws3.cell(r3, 4).value.date() == d1  # Date Found unchanged
    assert ws3.cell(r3, 5).value.date() == d3  # Last Seen updated
    assert ws3.cell(r3, 7).value == "1.1"  # Version bumped


def test_prune_missing_removes_deleted_files(tmp_path: Path) -> None:
    root = tmp_path / "root"
    root.mkdir()
    _write(root / "keep.pdf", b"%PDF-1.4\nkeep\n")
    _write(root / "gone.docx", b"docxcontent")

    out = tmp_path / "treasure_map.xlsx"
    d1 = date(2025, 12, 18)
    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=d1)

    # Delete a file and re-run WITHOUT prune: row should remain
    (root / "gone.docx").unlink()
    d2 = date(2025, 12, 19)
    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=d2, prune_missing=False)

    wb = load_workbook(out)
    ws = wb[MAIN_SHEET_NAME]
    locs = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}  # Location col 8
    assert "gone.docx" in locs
    assert "keep.pdf" in locs

    # Re-run WITH prune: deleted file should be removed
    d3 = date(2025, 12, 20)
    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=d3, prune_missing=True)

    wb2 = load_workbook(out)
    ws2 = wb2[MAIN_SHEET_NAME]
    locs2 = {ws2.cell(r, 8).value for r in range(2, ws2.max_row + 1)}  # Location col 8
    assert "gone.docx" not in locs2
    assert "keep.pdf" in locs2


def test_treasureignore_prunes_ignored_files(tmp_path: Path) -> None:
    root = tmp_path / "root"
    root.mkdir()

    # Create a document and generate the map
    _write(root / "keep.pdf", b"keep")
    _write(root / "ignore.pdf", b"ignore")

    out = tmp_path / "treasure_map.xlsx"
    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 18))

    wb = load_workbook(out)
    ws = wb[MAIN_SHEET_NAME]
    locs = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}  # Location col 8
    assert "keep.pdf" in locs
    assert "ignore.pdf" in locs

    # Introduce ignore rule and re-run WITHOUT prune_missing.
    # Ignored files should be removed from the map.
    (root / ".treasureignore").write_text("ignore.pdf\n", encoding="utf-8")

    create_or_update_treasure_map(root_dir=root, output_xlsx=out, today=date(2025, 12, 19), prune_missing=False)

    wb2 = load_workbook(out)
    ws2 = wb2[MAIN_SHEET_NAME]
    locs2 = {ws2.cell(r, 8).value for r in range(2, ws2.max_row + 1)}  # Location col 8
    assert "keep.pdf" in locs2
    assert "ignore.pdf" not in locs2
