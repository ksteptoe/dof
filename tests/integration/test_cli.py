from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from click.testing import CliRunner
from openpyxl import load_workbook

from dof import api
from dof.cli import BROKEN_LINKS_EXIT_CODE, cli

pytestmark = pytest.mark.integration


def _write(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def _headers(ws) -> dict:
    return {c.value: c.column for c in ws[1] if c.value}


def _locations(out: Path) -> set:
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    col = _headers(ws)["Location"]
    return {ws.cell(r, col).value for r in range(2, ws.max_row + 1)}


def test_cli_creates_output(tmp_path: Path) -> None:
    root = tmp_path / "root"
    root.mkdir()
    (root / "note.txt").write_text("hi", encoding="utf-8")

    out = tmp_path / "out.xlsx"
    runner = CliRunner()
    res = runner.invoke(cli, ["-d", str(root), "-o", str(out)])
    assert res.exit_code == 0
    assert out.exists()


def test_cli_keep_missing_flag(tmp_path: Path) -> None:
    """--keep-missing preserves deleted rows, and surfaces them as broken links.

    A retained row for a file that no longer exists is by definition a broken
    link, so the run exits ``BROKEN_LINKS_EXIT_CODE`` rather than 0.
    """
    root = tmp_path / "root"
    root.mkdir()
    (root / "a.pdf").write_bytes(b"%PDF-1.4\nA\n")
    (root / "b.docx").write_bytes(b"BBB")

    out = tmp_path / "out.xlsx"
    runner = CliRunner()
    res1 = runner.invoke(cli, ["-d", str(root), "-o", str(out)])
    assert res1.exit_code == 0

    # Remove one file - default behavior should prune it
    (root / "b.docx").unlink()
    res2 = runner.invoke(cli, ["-d", str(root), "-o", str(out)])
    assert res2.exit_code == 0

    assert "b.docx" not in _locations(out)  # Should be pruned by default

    # Re-add the file and run again
    (root / "b.docx").write_bytes(b"BBB")
    res3 = runner.invoke(cli, ["-d", str(root), "-o", str(out)])
    assert res3.exit_code == 0

    # Remove again but use --keep-missing
    (root / "b.docx").unlink()
    res4 = runner.invoke(cli, ["-d", str(root), "-o", str(out), "--keep-missing"])
    assert res4.exit_code == BROKEN_LINKS_EXIT_CODE
    assert "Broken links:" in res4.output
    assert "! b.docx" in res4.output

    assert "b.docx" in _locations(out)  # Should be kept with --keep-missing

    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = _headers(ws)
    statuses = {
        ws.cell(r, headers["Location"]).value: ws.cell(r, headers["Status"]).value for r in range(2, ws.max_row + 1)
    }
    assert statuses["b.docx"] == api.STATUS_BROKEN
    assert statuses["a.pdf"] == api.STATUS_OK


def test_cli_keep_missing_with_no_fail_on_broken_exits_zero(tmp_path: Path) -> None:
    """C3: the escape hatch keeps the exit code at 0 but still lists the broken rows."""
    root = tmp_path / "root"
    root.mkdir()
    (root / "a.pdf").write_bytes(b"%PDF-1.4\nA\n")
    (root / "b.docx").write_bytes(b"BBB")

    out = tmp_path / "out.xlsx"
    runner = CliRunner()
    assert runner.invoke(cli, ["-d", str(root), "-o", str(out)]).exit_code == 0

    (root / "b.docx").unlink()
    res = runner.invoke(cli, ["-d", str(root), "-o", str(out), "--keep-missing", "--no-fail-on-broken"])

    assert res.exit_code == 0
    assert "Broken links:" in res.output
    assert "! b.docx" in res.output
    assert "b.docx" in _locations(out)


def test_c1_dry_run_prints_moved_section(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "out.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")

    runner = CliRunner()
    assert runner.invoke(cli, ["-d", str(root), "-o", str(out)]).exit_code == 0

    (root / "b").mkdir()
    (root / "a" / "doc.pdf").rename(root / "b" / "doc.pdf")

    res = runner.invoke(cli, ["-d", str(root), "-o", str(out), "--dry-run"])

    assert res.exit_code == 0
    assert "Moved files:" in res.output
    assert "> a/doc.pdf -> b/doc.pdf" in res.output
    assert "Broken links:" not in res.output


def test_c4_clean_scan_prints_no_moved_or_broken_sections(tmp_path: Path) -> None:
    root = tmp_path / "root"
    out = tmp_path / "out.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")

    runner = CliRunner()
    assert runner.invoke(cli, ["-d", str(root), "-o", str(out)]).exit_code == 0
    res = runner.invoke(cli, ["-d", str(root), "-o", str(out)])

    assert res.exit_code == 0
    assert "Moved files:" not in res.output
    assert "Broken links:" not in res.output


@pytest.mark.parametrize("fmt", ["json", "csv"])
def test_c5_json_and_csv_exports_include_new_columns(tmp_path: Path, fmt: str) -> None:
    root = tmp_path / "root"
    out = tmp_path / "out.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")

    runner = CliRunner()
    res = runner.invoke(cli, ["-d", str(root), "-o", str(out), "--format", fmt])
    assert res.exit_code == 0

    written = out.with_suffix("." + fmt)
    assert written.exists()
    if fmt == "json":
        entry = json.loads(written.read_text(encoding="utf-8"))["treasure_map"][0]
        assert list(entry) == api.REQUIRED_COLUMNS
    else:
        with written.open(encoding="utf-8", newline="") as fh:
            header = next(csv.reader(fh))
        assert header == api.REQUIRED_COLUMNS


def _move_doc(root: Path) -> None:
    """Move ``a/doc.pdf`` to ``b/doc.pdf`` under *root*, byte-for-byte unchanged."""
    (root / "b").mkdir(exist_ok=True)
    (root / "a" / "doc.pdf").rename(root / "b" / "doc.pdf")


def _set_cell(out: Path, location: str, column: str, value) -> None:
    """Write *value* into *column* of the row for *location* in the workbook."""
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = _headers(ws)
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, headers["Location"]).value == location:
            ws.cell(r, headers[column], value)
    wb.save(out)


def test_c6_no_detect_moves_reports_new_plus_deleted(tmp_path: Path) -> None:
    """--no-detect-moves restores the pre-feature New/Deleted reporting."""
    root = tmp_path / "root"
    out = tmp_path / "out.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")

    runner = CliRunner()
    assert runner.invoke(cli, ["-d", str(root), "-o", str(out)]).exit_code == 0
    _move_doc(root)

    res = runner.invoke(cli, ["-d", str(root), "-o", str(out), "--dry-run", "--no-detect-moves"])

    assert res.exit_code == 0
    assert "  New:       1" in res.output
    assert "  Deleted:   1" in res.output
    assert "Moved:" not in res.output
    assert "Moved files:" not in res.output
    assert "+ b/doc.pdf" in res.output
    assert "- a/doc.pdf" in res.output


def test_c7_move_detection_is_on_by_default(tmp_path: Path) -> None:
    """The same scenario without the flag is reported as a move, not new + deleted."""
    root = tmp_path / "root"
    out = tmp_path / "out.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")

    runner = CliRunner()
    assert runner.invoke(cli, ["-d", str(root), "-o", str(out)]).exit_code == 0
    _move_doc(root)

    res = runner.invoke(cli, ["-d", str(root), "-o", str(out), "--dry-run"])

    assert res.exit_code == 0
    assert "  Moved:     1" in res.output
    assert "  New:       0" in res.output
    assert "Deleted:" not in res.output
    assert "> a/doc.pdf -> b/doc.pdf" in res.output


def test_c8_no_detect_moves_does_not_carry_row_history_over(tmp_path: Path) -> None:
    """With the flag the moved file lands as a brand-new row: no description, no history."""
    root = tmp_path / "root"
    out = tmp_path / "out.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")

    runner = CliRunner()
    assert runner.invoke(cli, ["-d", str(root), "-o", str(out)]).exit_code == 0
    _set_cell(out, "a/doc.pdf", "Description", "written by a user")
    _set_cell(out, "a/doc.pdf", "Date Found", "2001-01-01")

    _move_doc(root)
    res = runner.invoke(cli, ["-d", str(root), "-o", str(out), "--no-detect-moves"])
    assert res.exit_code == 0

    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = _headers(ws)
    assert _locations(out) == {"b/doc.pdf"}
    assert ws.cell(2, headers["Description"]).value in ("", None)
    assert ws.cell(2, headers["Version"]).value == "1.0"
    assert str(ws.cell(2, headers["Date Found"]).value) != "2001-01-01"


def test_bc1_workbook_from_the_previous_release_upgrades_in_place(tmp_path: Path) -> None:
    """BC1: a pre-move-tracking workbook gains the new columns without losing data."""
    root = tmp_path / "root"
    out = tmp_path / "out.xlsx"
    _write(root / "a" / "doc.pdf", b"content A")

    runner = CliRunner()
    assert runner.invoke(cli, ["-d", str(root), "-o", str(out)]).exit_code == 0

    # Downgrade the workbook to the previously released schema: eight columns on
    # the main sheet, two on the meta sheet.
    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    headers = _headers(ws)
    ws.cell(2, headers["Description"], "written by a user")
    for name in ("Previous Location", "Status"):
        ws.delete_cols(headers[name])
        headers = _headers(ws)
    meta_ws = wb[api.META_SHEET_NAME]
    meta_ws.delete_cols(_headers(meta_ws)["Size"])
    wb.save(out)

    old_headers = [c.value for c in load_workbook(out)[api.MAIN_SHEET_NAME][1]]
    assert old_headers == api.REQUIRED_COLUMNS[:8]

    res = runner.invoke(cli, ["-d", str(root), "-o", str(out)])
    assert res.exit_code == 0

    wb = load_workbook(out)
    ws = wb[api.MAIN_SHEET_NAME]
    assert [c.value for c in ws[1]] == api.REQUIRED_COLUMNS
    headers = _headers(ws)
    assert ws.cell(2, headers["Location"]).value == "a/doc.pdf"
    assert ws.cell(2, headers["Description"]).value == "written by a user"
    assert ws.cell(2, headers["Version"]).value == "1.0"
    assert ws.cell(2, headers["Status"]).value == api.STATUS_OK
    assert "Size" in _headers(wb[api.META_SHEET_NAME])
