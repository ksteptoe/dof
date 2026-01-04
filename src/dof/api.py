"""DOF API: build and maintain an Excel "treasure map" of documents in a directory tree.

Key behaviour:
- Recursively scan a directory for common document types.
- Write/update an Excel workbook (default: treasure_map.xlsx) with:
  File Name, File Type, Description, Date Found, Last Seen, Link, Version, Location
- If the workbook already exists:
  - identical file -> update Last Seen only
  - content changed -> increment Version, update Last Seen (Date Found remains first-seen)
  - file deleted -> keep row (unless --prune-missing), Last Seen frozen at last scan
  - file ignored via .treasureignore -> remove from map
"""

from __future__ import annotations

import csv
import hashlib
import json
import logging
import os
import sys
import tempfile
import urllib.parse
from dataclasses import dataclass, field
from datetime import date
from enum import Enum
from pathlib import Path, PurePosixPath
from typing import Callable, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from dof import __version__

_logger = logging.getLogger(__name__)

MAIN_SHEET_NAME = "treasure_map"
META_SHEET_NAME = "_dof_meta"  # hidden; stores fingerprints so we can detect *any* change


# A pragmatic "document" allowlist. Expand as needed.
DEFAULT_DOCUMENT_SUFFIXES = {
    # Office
    ".doc",
    ".docx",
    ".dot",
    ".dotx",
    ".rtf",
    ".xls",
    ".xlsx",
    ".xlsm",
    ".xlsb",
    ".xlt",
    ".xltx",
    ".xltm",
    ".ppt",
    ".pptx",
    ".pptm",
    ".pot",
    ".potx",
    # Text
    ".txt",
    ".text",
    ".md",
    ".rst",
    ".csv",
    ".tsv",
    ".yaml",
    ".yml",
    ".json",
    ".xml",
    ".toml",
    ".ini",
    # PDF
    ".pdf",
    # Other common docs
    ".odt",
    ".ods",
    ".odp",
    ".pages",
    ".numbers",
    ".key",
    ".epub",
    ".mobi",
    ".tex",
}


FILE_TYPE_BY_SUFFIX = {
    # PDF
    ".pdf": "PDF",
    # Text
    ".txt": "Text",
    ".text": "Text",
    ".md": "Markdown",
    ".rst": "reStructuredText",
    ".csv": "CSV",
    ".tsv": "TSV",
    ".rtf": "Rich Text",
    # Microsoft Office
    ".doc": "Word",
    ".docx": "Word",
    ".dot": "Word Template",
    ".dotx": "Word Template",
    ".xls": "Excel",
    ".xlsx": "Excel",
    ".xlsm": "Excel",
    ".xlsb": "Excel",
    ".xlt": "Excel Template",
    ".xltx": "Excel Template",
    ".xltm": "Excel Template",
    ".ppt": "PowerPoint",
    ".pptx": "PowerPoint",
    ".pptm": "PowerPoint",
    ".pot": "PowerPoint Template",
    ".potx": "PowerPoint Template",
    # OpenDocument
    ".odt": "OpenDocument Text",
    ".ods": "OpenDocument Spreadsheet",
    ".odp": "OpenDocument Presentation",
    # Apple iWork
    ".pages": "Pages",
    ".numbers": "Numbers",
    ".key": "Keynote",
    # eBooks
    ".epub": "EPUB",
    ".mobi": "Kindle",
    # Data/Config
    ".yaml": "YAML",
    ".yml": "YAML",
    ".json": "JSON",
    ".xml": "XML",
    ".toml": "TOML",
    ".ini": "INI",
    # Other
    ".tex": "LaTeX",
}


REQUIRED_COLUMNS = [
    "File Name",
    "File Type",
    "Description",
    "Date Found",
    "Last Seen",
    "Link",
    "Version",
    "Location",
]


class OutputFormat(Enum):
    XLSX = "xlsx"
    JSON = "json"
    CSV = "csv"


class ChangeType(Enum):
    NEW = "new"
    UPDATED = "updated"
    UNCHANGED = "unchanged"
    DELETED = "deleted"
    IGNORED = "ignored"


@dataclass
class FileChange:
    """Tracks a change to a single file for dry-run reporting."""

    location: str
    change_type: ChangeType
    old_version: Optional[str] = None
    new_version: Optional[str] = None


@dataclass
class ScanResult:
    """Result of a treasure map scan, used for dry-run and reporting."""

    total_found: int = 0
    new_files: List[str] = field(default_factory=list)
    updated_files: List[str] = field(default_factory=list)
    unchanged_files: List[str] = field(default_factory=list)
    deleted_files: List[str] = field(default_factory=list)
    ignored_files: List[str] = field(default_factory=list)
    changes: List[FileChange] = field(default_factory=list)

    def summary(self) -> str:
        """Return a human-readable summary of changes."""
        lines = [
            f"Total documents found: {self.total_found}",
            f"  New:       {len(self.new_files)}",
            f"  Updated:   {len(self.updated_files)}",
            f"  Unchanged: {len(self.unchanged_files)}",
        ]
        if self.deleted_files:
            lines.append(f"  Deleted:   {len(self.deleted_files)}")
        if self.ignored_files:
            lines.append(f"  Ignored:   {len(self.ignored_files)}")
        return "\n".join(lines)


@dataclass(frozen=True)
class FoundFile:
    abs_path: Path
    rel_location: str  # relative to root, POSIX-style
    filename: str
    suffix: str
    file_type: str
    sha256: Optional[str]


def _safe_save_workbook(wb, dest: Path) -> Path:
    """Save workbook safely.

    Writes to a temp file first, then attempts atomic replace.
    If destination is locked (e.g., open in Excel / OneDrive lock), writes to *.NEW.xlsx.
    """
    dest = dest.resolve()
    dest.parent.mkdir(parents=True, exist_ok=True)

    fd, tmp_name = tempfile.mkstemp(prefix=dest.stem + ".", suffix=".tmp.xlsx", dir=str(dest.parent))
    os.close(fd)
    tmp_path = Path(tmp_name)

    try:
        wb.save(tmp_path)

        try:
            tmp_path.replace(dest)  # atomic on same filesystem
            return dest
        except PermissionError:
            alt = dest.with_name(dest.stem + ".NEW" + dest.suffix)
            tmp_path.replace(alt)
            return alt
    finally:
        # Cleanup if anything went wrong and tmp still exists
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def setup_logging(loglevel: Optional[int]) -> None:
    """Setup basic logging."""
    if loglevel is None:
        loglevel = logging.WARNING
    logformat = "[%(asctime)s] %(levelname)s:%(name)s:%(message)s"
    logging.basicConfig(level=loglevel, stream=sys.stdout, format=logformat, datefmt="%Y-%m-%d %H:%M:%S")


def _posix_relpath(path: Path, root: Path) -> str:
    rel = path.resolve().relative_to(root.resolve())
    return rel.as_posix()


def _is_document(path: Path, suffixes: Iterable[str]) -> bool:
    return path.is_file() and path.suffix.lower() in suffixes


def _infer_file_type(suffix: str) -> str:
    s = suffix.lower()
    if s in FILE_TYPE_BY_SUFFIX:
        return FILE_TYPE_BY_SUFFIX[s]
    # fallback: ".foo" -> "FOO"
    return s.lstrip(".").upper() if s else "UNKNOWN"


def _sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _safe_sha256_file(path: Path) -> Optional[str]:
    """Best-effort SHA-256.

    OneDrive/Excel can temporarily lock files or expose cloud placeholders that
    raise PermissionError/OSError. In those cases we return None so the scan
    can continue without incorrectly bumping versions.
    """
    try:
        return _sha256_file(path)
    except (PermissionError, OSError):
        return None


def _hash_changed(old: Optional[str], new: Optional[str]) -> bool:
    """Return True only when we can *prove* a content change.

    - If we can't read/hash the file now (new is None), we do not claim change.
    - If we couldn't hash it previously (old is None) but can now, treat as no change
      (metadata improvement only).
    """
    if new is None:
        return False
    if old is None:
        return False
    return old != new


@dataclass(frozen=True)
class IgnoreRule:
    pattern: str
    negated: bool = False
    dir_only: bool = False
    root_anchored: bool = False


def _load_treasureignore(root_dir: Path) -> Optional[List[IgnoreRule]]:
    """Load .treasureignore from root_dir (gitignore-ish patterns).

    This is a pragmatic subset:
    - blank lines and lines starting with # are ignored
    - negation with leading ! is supported (last match wins)
    - patterns with no / match anywhere (we also try **/pattern)
    - patterns starting with / match only at root level
    - patterns ending with / ignore that directory and everything under it
    - ** is supported via PurePosixPath.match
    """
    ignore_path = root_dir / ".treasureignore"
    if not ignore_path.exists() or not ignore_path.is_file():
        return None

    rules: List[IgnoreRule] = []
    try:
        for raw in ignore_path.read_text(encoding="utf-8", errors="replace").splitlines():
            line = raw.strip()
            if not line:
                continue
            if line.startswith("#"):
                continue

            neg = line.startswith("!")
            if neg:
                line = line[1:].strip()
                if not line:
                    continue

            # root-anchored patterns start with /
            root_anchored = line.startswith("/")
            if root_anchored:
                line = line[1:]

            dir_only = line.endswith("/")
            if dir_only:
                line = line[:-1].strip()
                if not line:
                    continue

            rules.append(IgnoreRule(pattern=line, negated=neg, dir_only=dir_only, root_anchored=root_anchored))
    except OSError:
        return None

    return rules or None


def _rule_matches(rel_posix: str, rule: IgnoreRule) -> bool:
    p = PurePosixPath(rel_posix)
    pat = rule.pattern

    # Root-anchored patterns only match at the root level (no parent directories)
    if rule.root_anchored:
        if rule.dir_only:
            # Root-anchored directory pattern: must match at root and be a directory
            return p.match(pat + "/**") or p.match(pat)
        # Root-anchored file/pattern: must match at root level (no / in rel_posix)
        return "/" not in rel_posix and p.match(pat)

    if rule.dir_only:
        # match any path under that directory; treat pattern as a path fragment
        # e.g. "tmp/" matches "tmp/a.pdf" and "a/tmp/b.pdf" when pattern has no "/".
        if "/" in pat:
            return p.match(pat + "/**") or p.match(pat)
        # directory name anywhere in the path
        for i in range(1, len(p.parts) + 1):
            prefix = PurePosixPath(*p.parts[:i])
            if prefix.match(pat):
                # ensure it's actually a directory boundary: prefix shorter than full path
                return True
        return False

    if "/" in pat:
        return p.match(pat)

    # basename-style patterns should match anywhere
    return p.match(pat) or p.match("**/" + pat)


def _is_ignored(rel_posix: str, rules: Optional[List[IgnoreRule]]) -> bool:
    if not rules:
        return False
    ignored = False
    for r in rules:
        if _rule_matches(rel_posix, r):
            ignored = not r.negated
    return ignored


def _build_sharepoint_url(base: Optional[str], rel_location_posix: str, abs_path: Path) -> str:
    """Create a hyperlink target.

    If base is provided, treat it as a SharePoint/OneDrive base URL and append rel path.
    Otherwise, fall back to a local file:// URI.
    """
    if base:
        # Preserve existing querystring/fragments on base; append path.
        # Use urllib for safe quoting of path segments.
        base = base.rstrip("/")
        # SharePoint URLs are typically already encoded; we encode the rel path safely.
        rel_parts = [urllib.parse.quote(p) for p in rel_location_posix.split("/")]
        return base + "/" + "/".join(rel_parts)
    return abs_path.resolve().as_uri()


def discover_documents(
    root_dir: Path,
    suffixes: Optional[Iterable[str]] = None,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> List[FoundFile]:
    """Recursively scan root_dir for document files.

    Respects an optional .treasureignore in root_dir, using gitignore-style patterns.
    Uses Path.walk() for efficient directory traversal (Python 3.12+).

    Args:
        root_dir: Directory to scan.
        suffixes: File extensions to include (defaults to DEFAULT_DOCUMENT_SUFFIXES).
        progress_callback: Optional callback called with each file path being processed.
    """
    suffixes_set = set(s.lower() for s in (suffixes or DEFAULT_DOCUMENT_SUFFIXES))
    ignore_rules = _load_treasureignore(root_dir)

    found: List[FoundFile] = []

    # Use Path.walk() (Python 3.12+) for more efficient traversal
    for dirpath, _dirnames, filenames in root_dir.walk():
        for filename in filenames:
            # Always ignore the ignore file itself
            if filename == ".treasureignore":
                continue

            p = dirpath / filename
            suffix = p.suffix.lower()
            if suffix not in suffixes_set:
                continue

            rel = _posix_relpath(p, root_dir)
            if _is_ignored(rel, ignore_rules):
                continue

            if progress_callback:
                progress_callback(rel)

            found.append(
                FoundFile(
                    abs_path=p,
                    rel_location=rel,
                    filename=filename,
                    suffix=suffix,
                    file_type=_infer_file_type(suffix),
                    sha256=_safe_sha256_file(p),
                )
            )

    # deterministic order helps tests and diffing
    found.sort(key=lambda x: x.rel_location.lower())
    return found


def _ensure_required_headers(ws: Worksheet) -> Dict[str, int]:
    """Return a mapping column_name -> 1-based column index, ensuring headers exist."""
    header_row = 1
    existing = [c.value for c in ws[header_row]]
    mapping: Dict[str, int] = {}
    for idx, name in enumerate(existing, start=1):
        if isinstance(name, str) and name.strip():
            mapping[name.strip()] = idx

    # Create headers if sheet is empty or missing required columns.
    if not mapping:
        for col_idx, name in enumerate(REQUIRED_COLUMNS, start=1):
            ws.cell(row=header_row, column=col_idx, value=name)
            mapping[name] = col_idx
    else:
        # Add any missing required columns to the right.
        max_col = ws.max_column
        for name in REQUIRED_COLUMNS:
            if name not in mapping:
                max_col += 1
                ws.cell(row=header_row, column=max_col, value=name)
                mapping[name] = max_col
    return mapping


def _style_header(ws: Worksheet, mapping: Dict[str, int]) -> None:
    font = Font(bold=True)
    for name, col in mapping.items():
        if name in REQUIRED_COLUMNS:
            ws.cell(row=1, column=col).font = font
    ws.freeze_panes = "A2"


def _autosize_columns(ws: Worksheet, mapping: Dict[str, int], max_width: int = 80) -> None:
    # crude autosize based on max string length in each column
    for name, col in mapping.items():
        if name not in REQUIRED_COLUMNS:
            continue
        max_len = len(name)
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            v_str = str(v)
            max_len = max(max_len, len(v_str))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, max_width)


def _load_or_create_workbook(output_xlsx: Path) -> Tuple[Workbook, Worksheet, Worksheet]:
    """Load an existing workbook if possible, otherwise create a new one.

    Returns (wb, main_ws, meta_ws).
    """
    if output_xlsx.exists():
        try:
            wb = load_workbook(output_xlsx)
        except PermissionError:
            _logger.warning(
                "Cannot open %s (locked/open in Excel). Cannot preserve existing Description values.",
                output_xlsx,
            )
            wb = Workbook()
            ws = wb.active
            ws.title = MAIN_SHEET_NAME
            meta_ws = wb.create_sheet(META_SHEET_NAME)
            meta_ws.sheet_state = "hidden"
            return wb, ws, meta_ws

        # Main sheet
        if MAIN_SHEET_NAME in wb.sheetnames:
            ws = wb[MAIN_SHEET_NAME]
        else:
            ws = wb.active
            ws.title = MAIN_SHEET_NAME

        # Meta sheet
        if META_SHEET_NAME in wb.sheetnames:
            meta_ws = wb[META_SHEET_NAME]
        else:
            meta_ws = wb.create_sheet(META_SHEET_NAME)
            meta_ws.sheet_state = "hidden"

        return wb, ws, meta_ws

    # New workbook
    wb = Workbook()
    ws = wb.active
    ws.title = MAIN_SHEET_NAME
    meta_ws = wb.create_sheet(META_SHEET_NAME)
    meta_ws.sheet_state = "hidden"
    return wb, ws, meta_ws


def _read_existing_rows(ws: Worksheet, mapping: Dict[str, int]) -> Dict[str, Dict[str, object]]:
    """Return existing rows keyed by Location (relative path)."""
    rows: Dict[str, Dict[str, object]] = {}
    loc_col = mapping["Location"]

    for r in range(2, ws.max_row + 1):
        loc = ws.cell(row=r, column=loc_col).value
        if not loc:
            continue
        loc_str = str(loc)
        row_data: Dict[str, object] = {"__row": r}
        for name in REQUIRED_COLUMNS:
            row_data[name] = ws.cell(row=r, column=mapping[name]).value
        # preserve hyperlink target (openpyxl stores it on the cell)
        link_cell = ws.cell(row=r, column=mapping["Link"])
        if link_cell.hyperlink:
            row_data["__link_target"] = link_cell.hyperlink.target
        rows[loc_str] = row_data
    return rows


def _meta_headers(meta_ws: Worksheet) -> Dict[str, int]:
    mapping = {}
    if meta_ws.max_row >= 1:
        existing = [c.value for c in meta_ws[1]]
        for idx, v in enumerate(existing, start=1):
            if isinstance(v, str) and v.strip():
                mapping[v.strip()] = idx
    if not mapping:
        meta_ws.cell(1, 1, "Location")
        meta_ws.cell(1, 2, "Sha256")
        mapping = {"Location": 1, "Sha256": 2}
    return mapping


def _read_meta(meta_ws: Worksheet) -> Dict[str, Optional[str]]:
    mapping = _meta_headers(meta_ws)
    out: Dict[str, Optional[str]] = {}
    for r in range(2, meta_ws.max_row + 1):
        loc = meta_ws.cell(r, mapping["Location"]).value
        sha = meta_ws.cell(r, mapping["Sha256"]).value
        if loc:
            out[str(loc)] = str(sha) if sha is not None else None
    return out


def _write_meta(meta_ws: Worksheet, meta: Dict[str, Optional[str]]) -> None:
    mapping = _meta_headers(meta_ws)
    # clear old
    if meta_ws.max_row > 1:
        meta_ws.delete_rows(2, meta_ws.max_row - 1)
    # write deterministic
    for i, loc in enumerate(sorted(meta.keys(), key=lambda s: s.lower()), start=2):
        meta_ws.cell(i, mapping["Location"], loc)
        meta_ws.cell(i, mapping["Sha256"], meta[loc])


def _parse_version(v: object) -> Tuple[int, int]:
    """Return (major, minor_tenths). Defaults to (1,0) if unparseable."""
    if v is None:
        return 1, 0
    s = str(v).strip()
    m = s.split(".")
    try:
        major = int(m[0])
        minor = int(m[1]) if len(m) > 1 else 0
        return major, minor
    except (ValueError, IndexError):
        return 1, 0


def _bump_version(v: object) -> str:
    major, minor = _parse_version(v)
    minor += 1
    return f"{major}.{minor}"


def _set_link_cell(cell: Cell, target: str, text: str) -> None:
    cell.value = text
    cell.hyperlink = target
    cell.style = "Hyperlink"


def _row_to_dict(row: Dict[str, object]) -> Dict[str, object]:
    """Convert internal row representation to a clean dict for JSON/CSV export."""
    result = {}
    for col in REQUIRED_COLUMNS:
        val = row.get(col)
        if col == "Link":
            # Extract URL from link dict
            if isinstance(val, dict):
                result[col] = val.get("target", "")
            else:
                result[col] = str(val) if val else ""
        elif isinstance(val, date):
            result[col] = val.isoformat()
        else:
            result[col] = val if val is not None else ""
    return result


def _write_json(output_path: Path, rows: Dict[str, Dict[str, object]]) -> Path:
    """Write treasure map data to JSON file."""
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    data = {"treasure_map": [_row_to_dict(rows[loc]) for loc in sorted(rows.keys(), key=lambda s: s.lower())]}

    with output_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    return output_path


def _write_csv(output_path: Path, rows: Dict[str, Dict[str, object]]) -> Path:
    """Write treasure map data to CSV file."""
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=REQUIRED_COLUMNS)
        writer.writeheader()
        for loc in sorted(rows.keys(), key=lambda s: s.lower()):
            writer.writerow(_row_to_dict(rows[loc]))

    return output_path


def create_or_update_treasure_map(
    *,
    root_dir: Path,
    output_xlsx: Path,
    sharepoint_base_url: Optional[str] = None,
    today: Optional[date] = None,
    suffixes: Optional[Iterable[str]] = None,
    prune_missing: bool = False,
    dry_run: bool = False,
    output_format: OutputFormat = OutputFormat.XLSX,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> Path | ScanResult:
    """Scan root_dir and create/update the treasure map.

    Args:
        root_dir: Directory to scan for documents.
        output_xlsx: Output file path (extension ignored for JSON/CSV formats).
        sharepoint_base_url: Optional SharePoint base URL for hyperlinks.
        today: Override today's date (for testing).
        suffixes: File extensions to include.
        prune_missing: If True, remove rows for files that no longer exist.
        dry_run: If True, compute changes but don't write files. Returns ScanResult.
        output_format: Output format (XLSX, JSON, or CSV).
        progress_callback: Optional callback for progress reporting.

    Returns:
        Path to written file, or ScanResult if dry_run=True.
    """
    today = today or date.today()
    root_dir = root_dir.resolve()
    output_xlsx = output_xlsx.resolve()

    _logger.info("DOF %s", __version__)
    _logger.info("Scanning %s", root_dir)

    found = discover_documents(root_dir, suffixes=suffixes, progress_callback=progress_callback)
    _logger.info("Found %d document(s)", len(found))

    # Track changes for dry-run reporting
    scan_result = ScanResult(total_found=len(found))

    # For JSON/CSV with no existing file, we don't need to load a workbook
    if output_format != OutputFormat.XLSX and not output_xlsx.exists():
        existing_rows: Dict[str, Dict[str, object]] = {}
        meta: Dict[str, Optional[str]] = {}
        wb = None
        ws = None
        meta_ws = None
        mapping = {col: i + 1 for i, col in enumerate(REQUIRED_COLUMNS)}
    else:
        wb, ws, meta_ws = _load_or_create_workbook(output_xlsx)
        mapping = _ensure_required_headers(ws)
        _style_header(ws, mapping)
        existing_rows = _read_existing_rows(ws, mapping)
        meta = _read_meta(meta_ws)

    # We'll build a new in-memory table of rows, preserving existing rows + appending new ones.
    updated_rows: Dict[str, Dict[str, object]] = dict(existing_rows)

    for f in found:
        loc = f.rel_location
        prev_sha = meta.get(loc)

        if loc in existing_rows:
            row = updated_rows[loc]
            old_version = str(row.get("Version", "1.0"))

            # Always update Last Seen for files that still exist in the scan
            row["Last Seen"] = today

            # identical (including both None) -> no change (except Last Seen)
            if prev_sha == f.sha256:
                scan_result.unchanged_files.append(loc)
                scan_result.changes.append(FileChange(loc, ChangeType.UNCHANGED, old_version, old_version))
                continue

            # changed (provable) -> bump Version only (Date Found is first-seen)
            if _hash_changed(prev_sha, f.sha256):
                new_version = _bump_version(row.get("Version"))
                row["Version"] = new_version
                meta[loc] = f.sha256
                scan_result.updated_files.append(loc)
                scan_result.changes.append(FileChange(loc, ChangeType.UPDATED, old_version, new_version))
                continue

            # Previously unreadable/unhashed but now readable -> record hash, no bump
            if prev_sha is None and f.sha256 is not None:
                meta[loc] = f.sha256
                scan_result.unchanged_files.append(loc)
                scan_result.changes.append(FileChange(loc, ChangeType.UNCHANGED, old_version, old_version))
                continue

            # Hashed before but unreadable now -> no change
            scan_result.unchanged_files.append(loc)
            scan_result.changes.append(FileChange(loc, ChangeType.UNCHANGED, old_version, old_version))
            continue

        # New file -> create a new row
        link_target = _build_sharepoint_url(sharepoint_base_url, loc, f.abs_path)
        updated_rows[loc] = {
            "File Name": f.filename,
            "File Type": f.file_type,
            "Description": "",
            "Date Found": today,
            "Last Seen": today,
            "Link": {"target": link_target, "text": f.filename},
            "Version": "1.0",
            "Location": loc,
        }
        meta[loc] = f.sha256
        scan_result.new_files.append(loc)
        scan_result.changes.append(FileChange(loc, ChangeType.NEW, None, "1.0"))

    # If a .treasureignore exists, treat ignored files as out-of-scope and remove them.
    ignore_rules_map = _load_treasureignore(root_dir)
    if ignore_rules_map:
        for loc in list(updated_rows.keys()):
            if _is_ignored(str(loc).replace("\\", "/"), ignore_rules_map):
                updated_rows.pop(loc, None)
                meta.pop(loc, None)
                scan_result.ignored_files.append(loc)
                scan_result.changes.append(FileChange(loc, ChangeType.IGNORED))

    if prune_missing:
        # Remove rows for files that no longer exist in the scanned tree.
        found_locs = {f.rel_location for f in found}
        for loc in list(updated_rows.keys()):
            if loc not in found_locs:
                updated_rows.pop(loc, None)
                meta.pop(loc, None)
                scan_result.deleted_files.append(loc)
                scan_result.changes.append(FileChange(loc, ChangeType.DELETED))

    # If dry run, return the scan result without writing
    if dry_run:
        _logger.info("Dry run - no files written")
        return scan_result

    # Write output based on format
    if output_format == OutputFormat.JSON:
        output_path = output_xlsx.with_suffix(".json")
        written = _write_json(output_path, updated_rows)
        _logger.info("Wrote %s", written)
        return written

    if output_format == OutputFormat.CSV:
        output_path = output_xlsx.with_suffix(".csv")
        written = _write_csv(output_path, updated_rows)
        _logger.info("Wrote %s", written)
        return written

    # XLSX format - need workbook
    if wb is None:
        wb, ws, meta_ws = _load_or_create_workbook(output_xlsx)
        mapping = _ensure_required_headers(ws)
        _style_header(ws, mapping)

    # Rewrite the main sheet (keeps it clean + deterministic)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Keep deterministic ordering by Location
    for row_idx, loc in enumerate(sorted(updated_rows.keys(), key=lambda s: s.lower()), start=2):
        row = updated_rows[loc]
        for col_name in REQUIRED_COLUMNS:
            c = ws.cell(row=row_idx, column=mapping[col_name])
            if col_name == "Link":
                link = row.get("Link")
                if isinstance(link, dict) and link.get("target"):
                    _set_link_cell(c, link["target"], link.get("text") or row.get("File Name", ""))
                else:
                    # Preserve an existing hyperlink if present and rule says "no other change"
                    existing = existing_rows.get(loc, {})
                    if "__link_target" in existing and existing.get("Link"):
                        _set_link_cell(c, existing["__link_target"], str(existing["Link"]))
                    else:
                        c.value = row.get("Link", "")
            else:
                c.value = row.get(col_name, "")

        for col_name in ("Date Found", "Last Seen"):
            dcell = ws.cell(row=row_idx, column=mapping[col_name])
            if isinstance(dcell.value, date):
                dcell.number_format = "dd/mm/yyyy"

    _autosize_columns(ws, mapping)
    _write_meta(meta_ws, meta)

    written = _safe_save_workbook(wb, output_xlsx)
    _logger.info("Wrote %s", written)

    if written != output_xlsx:
        _logger.warning(
            "Destination was locked (likely open in Excel). Wrote to %s instead.",
            written,
        )

    return written


def dof_api(
    loglevel: Optional[int],
    *,
    root_dir: Path,
    output_xlsx: Path,
    sharepoint_base_url: Optional[str] = None,
    prune_missing: bool = False,
    dry_run: bool = False,
    output_format: OutputFormat = OutputFormat.XLSX,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> Path | ScanResult:
    """CLI-friendly wrapper."""
    setup_logging(loglevel)
    return create_or_update_treasure_map(
        root_dir=root_dir,
        output_xlsx=output_xlsx,
        sharepoint_base_url=sharepoint_base_url,
        prune_missing=prune_missing,
        dry_run=dry_run,
        output_format=output_format,
        progress_callback=progress_callback,
    )
